import os
import sys
import re
import json
import math
import random
import string
import openpyxl
import pandas as pd

from os.path import expanduser
from time import sleep
from datetime import date, time, datetime
from PyQt5.QtCore import QThread, pyqtSignal
from openpyxl.styles import Alignment, Font, NamedStyle, PatternFill
from openpyxl import formatting, styles, Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule

class Report:

    def __init__(self, serial, output_path, nowTime, df_result, start_time, end_time, take_time, parent=None):

        self.output_path = output_path
        self.df_result = df_result
        self.serial = serial
        self.nowTime = nowTime
        self.start_time = start_time
        self.end_time = end_time
        self.take_time = take_time
        self.file_name = "result_"+self.serial+"("+self.nowTime+")"

    def html_report(self):

        self.html_body = '<tbody>'
        self.html_header = """
        <!doctype html>
        <html>
            <head>
                <title>Appium Test Report</title>
                <meta charset="utf-8">
                <meta name="viewport" content="width=device-width, initial-scale=1, shrink-to-fit=no">
                <style type='text/css'>
                table{
                    border: 1px solid rgba(225,194,207,100);
                    border-radius:5px;
                    width: 100%;
                    table-layout:fixed;
                }
                td,th{
                    border: 1px solid rgba(225,194,207,100);
                }

                .testcase, .ele-value, .assert-value, .event-desc, .assert-desc{
                    word-break:break-all;
                }
                .event-type, .event-api, .assert-type, .assert-api, .assert-result{
                    word-break:break-all;
                }

                .elapsed{
                    word-break:break-all;
                }

                #times{

                }
                .test-result-table-header-cell {
                    background-color:rgba(133,47,65,52);
                    color:rgba(255,255,255,100);
                    font-size:15px;
                }

                </style>
            </head>
            <body>
                <h1 class='test-results-header'>
                    Device Name : """+self.serial+"""
                </h1>
                <table class='test-result-table' cellspacing='0'>
                    <colgroup>
                        <col style="width:15%";/>
                        <col style="width:4.5%";/>
                        <col style="width:4.5%";/>
                        <col style="width:22.5%";/>
                        <col style="width:4.5%";/>
                        <col style="width:4.5%";/>
                        <col style="width:12%";/>
                        <col style="width:4%";/>
                        <col style="width:13%";/>
                        <col style="width:13%";/>
                        <col style="width:5%";/>
                    </colgroup>
                    <thead>
                        <tr>
                            <div id=times>
                                <h4> Start_Time: """+self.start_time+"""&nbsp&nbspEnd_Time: """+self.end_time+"""&nbsp&nbspElapesd: """+str(round(float(self.take_time)/60,3))+"""</h4>
                            </div>
                        </tr>
                        <tr>
                            <th class='test-result-table-header-cell testcase'>
                                TestCase
                            </th>
                            <th class='test-result-table-header-cell event-api'>
                                Event_API
                            </th>
                            <th class='test-result-table-header-cell ele-type'>
                                Ele_Type
                            </th>
                            <th class='test-result-table-header-cell ele-value'>
                                Ele_Value
                            </th>
                            <th class='test-result-table-header-cell assert-api'>
                                Assert_API
                            </th>
                            <th class='test-result-table-header-cell assert-type'>
                                Assert_Type
                            </th>
                            <th class='test-result-table-header-cell assert-value'>
                                Assert_Value
                            </th>
                            <th class='test-result-table-header-cell assert-result'>
                                Result
                            </th>
                            <th class='test-result-table-header-cell event-desc'>
                                Event-Desc
                            </th>
                            <th class='test-result-table-header-cell assert-desc'>
                                Assert_Desc
                            </th>
                            <th class='test-result-table-header-cell elapsed'>
                                Elapsed
                            </th>
                        </tr>
                    </thead>
        """
        for i in range(self.df_result.shape[0]):
            self.html_tr = """
            <tr class='test-result-step-row test-result-step-row-altone'>
                <td class='test-result-step-description-cell testcase'>"""+str(self.df_result.at[i,"TestCase"])+"""
                </td>
                <td class='test-result-step-description-cell event-api'>"""+str(self.df_result.at[i,"Event_API"])+"""
                </td>
                </td>
                <td class='test-result-step-description-cell ele-type'>"""+str(self.df_result.at[i,"Ele_Type"])+"""
                </td>
                </td>
                <td class='test-result-step-description-cell ele-value'>"""+str(self.df_result.at[i,"Ele_Value"])+"""
                </td>
                </td>
                <td class='test-result-step-description-cell assert-api'>"""+str(self.df_result.at[i,"Assert_API"])+"""
                </td>
                </td>
                <td class='test-result-step-description-cell assert-type'>"""+str(self.df_result.at[i,"Assert_Type"])+"""
                </td>
                </td>
                <td class='test-result-step-description-cell assert-value'>"""+str(self.df_result.at[i,"Assert_Value"])+"""
                </td>
                <td class='test-result-step-description-cell result-take assert-result'>"""+str(self.df_result.at[i,"Result"])+"""
                </td>
                <td class='test-result-step-description-cell event-desc'>"""+str(self.df_result.at[i,"Event_Desc"])+"""
                </td>
                <td class='test-result-step-description-cell assert-desc'>"""+str(self.df_result.at[i,"Assert_Desc"])+"""
                </td>
                <td class='test-result-step-description-cell elapsed'>"""+str(self.df_result.at[i,"Elapsed"])+"""
                </td>
            </tr>
            """
            self.html_body = self.html_body+self.html_tr

        self.html_script = """
        <script type="text/javascript">
            window.onload = function(){
                var results = document.getElementsByClassName('result-take');
                for(var i=0; i<results.length; i++){
                    if(results[i].innerText == 'PASS'){
                        results[i].style.color = "white";
                        results[i].style.backgroundColor = 'rgba(117,120,255,100)';
                    }else{
                        results[i].style.color = "white";
                        results[i].style.backgroundColor = 'rgba(201,0,0,83)';
                    }
                }
            };
        </script>
        """
        self.html_final = self.html_header + self.html_body + "</tbody></table></body>"+self.html_script+"</html>"
        with open(self.output_path + self.file_name+".txt", "w", encoding='utf8') as fw:
            fw.write(self.html_final)

    def excel_report(self):

        self.writer = pd.ExcelWriter(self.output_path + self.file_name+".xlsx", options={'encoding':'utf-8'})
        self.df_result.to_excel(self.writer, sheet_name=self.serial, index=False)
        self.writer.save()
        #Excel Options deploy part]
        self.rows = self.df_result.shape[0]+3
        # self.date_style = NamedStyle(name="datetime", number_format="YYYY-MM-DD")
        self.wb = openpyxl.load_workbook(self.output_path + self.file_name+".xlsx")
        self.sheet = self.wb[self.serial]
        # set title insert row 2
        for i in range(2):
            self.sheet.insert_rows(1)

        self.list_cellName = ['A','B','C','D','E','F','G','H','I','J','K']
        self.list_cellWidth = [34.13, 9.88, 9.88, 34.13, 9.88, 9.88, 34.13, 8.38, 25, 25, 16.88]
        #set title excel_report
        self.sheet.merge_cells('A1:E1')
        self.sheet.merge_cells('A2:E2')
        self.sheet['A1'] = "Device Name: "+self.serial
        self.sheet['A1'].font = Font(size=16, bold=True)
        self.sheet['A1'].alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")
        self.sheet['A2'] = "Start: "+self.start_time+"   End: "+self.end_time+"   Elaped: "+str(round(float(self.take_time)/60,3))
        self.sheet['A2'].font = Font(size=14)
        self.sheet['A2'].alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")

        # cell width adjusting
        for i in range(len(self.list_cellName)):
            self.sheet.column_dimensions[self.list_cellName[i]].width = self.list_cellWidth[i]
        #set cell filtering
        self.sheet.auto_filter.ref = "A3:K"+str(self.rows)
        #set cell styles
        for mCell in self.sheet["A3:K"+str(self.rows)]:
            for cell in mCell:
                # cell.style = self.date_style
                if cell.row == 3:
                    cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
                    cell.font = Font(size=10, bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color="852F41", end_color="852F41", fill_type = "solid")
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                else:
                    cell.alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")
                    cell.font = Font(size=10)
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        # set result Style by which 'PASS' or 'FAIL'
        for mCell in self.sheet['H4:H'+str(self.rows)]:
            for cell in mCell:
                if cell.value == "PASS":
                    cell.fill = PatternFill(start_color="7578FF", end_color="7578FF",fill_type = "solid")
                else:
                    cell.fill = PatternFill(start_color="D40000", end_color="D40000", fill_type = "solid")
                cell.font = Font(size=10, bold=True, color='FFFFFF')

        # self.white_font = styles.Font(size=10, bold=True, color='FFFFFF')
        # self.red_fill = styles.PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        # self.blue_fill = styles.PatternFill(start_color='8DB4E2', end_color='8DB4E2', fill_type='solid')
        # self.sheet.conditional_formatting.add('H2:H'+str(self.rows), formatting.rule.CellIsRule(operator='equal', formula=['PASS'], stopIfTrue=True, fill=self.blue_fill, font=self.white_font))
        # self.sheet.conditional_formatting.add('H2:H'+str(self.rows), formatting.rule.CellIsRule(operator='equal', formula=['FAIL'], stopIfTrue=True, fill=self.red_fill, font=self.white_font))
        self.wb.save(self.output_path + self.file_name+".xlsx")
