# -*- coding: utf-8 -*-
"""
Created on Mon May 21 13:36:12 2018
@author: TestEnC hanrim lee

"""
import os
import sys
import re
import pandas as pd
import ctypes
from datetime import date, time, datetime
from os.path import expanduser

import openpyxl
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Font, NamedStyle, PatternFill, Color, Border, Side
# from openpyxl.styles.colors import Color
# from openpyxl.styles.fills import PatternFill,Fill
# from openpyxl.chart import (AreaChart,Reference,Series,LineChart)

# #color table
# STD_INPUT_HANDLE   = -10
# STD_OUTPUT_HANDLE  = -11
# STD_ERROR_HANDLE   = -12
#
# FOREGROUND_BLACK     = 0x00
# FOREGROUND_BLUE      = 0x01 # text color contains blue.
# FOREGROUND_GREEN     = 0x02 # text color contains green.
# FOREGROUND_RED       = 0x04 # text color contains red.
# FOREGROUND_INTENSITY = 0x08 # text color is intensified.
# BACKGROUND_BLUE      = 0x10 # background color contains blue.
# BACKGROUND_GREEN     = 0x20 # background color contains green.
# BACKGROUND_RED       = 0x40 # background color contains red.
# BACKGROUND_INTENSITY = 0x80 # background color is intensified.
# std_out_handle = ctypes.windll.kernel32.GetStdHandle(STD_OUTPUT_HANDLE)
# #set color number 1 => Blue
# #set color number 2 => Green
# #set color number 3 => Navy
# #set color number 4 => Red
# #set color number 5 => Purple
# #set color number 6 => Yellow
# #set color number 7 => White(reset)
# #set-text Background color number 17 => Blue Back (plus 16 to text corlor number)
# #set-text Background color number 18 => Green Back
# #Function collection
#
# def divString(display):
#     strToday = datetime.today().strftime("%Y-%m-%d %H:%M:%S")
#     print("\n\n"+strToday+" ==> ####################{}####################\n\n".format(display))
#
# def ingString(display):
#     strToday = datetime.today().strftime("%Y-%m-%d %H:%M:%S")
#     print("\n"+strToday+" ==> {}...".format(display))

# def setPrintColor(color,text):
#     strToday = datetime.today().strftime("%Y-%m-%d %H:%M:%S")
#     color_dic = {"blue":1, "red":4, "navy":3, "yellow":6, "purple":5, "green":2, "white":7}
#     try:
#         setColor(color_dic[color])
#         print("\n-- {}".format(text))
#         resetColor()
#     except:
#         setColor(1)
#         print("\n"+strToday+" ==> {}".format(text))
# def setColor(color, handle= std_out_handle):
#     bool = ctypes.windll.kernel32.SetConsoleTextAttribute(handle, color)
#     return bool

# def resetColor():
#     setColor(7)

def printText(text):
    now = datetime.today().strftime('%Y-%m-%d %H:%M:%S')
    print(now + ":\n".format(text))

def removeString(text):

    tempText = text.replace("[", "")
    tempText = tempText.replace("]", "")
    tempText = tempText.replace("-", "")
    tempText = tempText.replace("/", "")
    tempText = tempText.replace("=", "")
    tempText = tempText.replace("<", "")
    tempText = tempText.replace(">", "")
    tempText = tempText.replace("#", "")
    tempText = tempText.replace("*", "")
    # parse = sub(r"[-=.#/?:$}\\><\']","",text)
    # parse  = re.sub("[0-9]:| [0-9].","",parse)
    return tempText

def find_between(s, first, last):
    try:
        start = s.index(first)+len(first)
        end = s.index(last, start)
        return s[start:end]
    except ValueError:
        return ""

def find_between_r(s, first, last ):
    try:
        start = s.rindex(first) + len(first)
        end = s.rindex(last, start)
        return s[start:end]
    except ValueError:
        return ""

def insert_row(idx, df, df_insert):
    dfA = df.iloc[:idx, ]
    dfB = df.iloc[idx:, ]

    cdrRows = dfA.shape[0]
    cdrColumns_list = list(dfA)
    insertValuse_list = list(df_insert)
    # dfA.append(pd.Series([np.nan]), ignore_index = True)
    for i in range(len(cdrColumns_list)):
        dfA[cdrColumns_list[i]] = dfA[cdrColumns_list[i]].astype(str)
        dfA.at[cdrRows+1, cdrColumns_list[i]] = str(insertValuse_list[i])
    df = dfA.append(dfB).reset_index(drop=True)
    return df

def setNewCol(col_list,df):
    for item in col_list:
        df[item] = ""
    return df

def tNumberFormat(list_tNum):

    list_return = list_tNum
    for idx, tNum in enumerate(list_return):
        if len(tNum) < 4:
            list_return[idx] = "0" + tNum
        pass
    return list_return

def checkConfigData(list):

    dict_item = {0:'시험명', 1:'시험번호', 2:'시작시간', 3:'종료시간', 4:'포트', 5:'변환타입'}
    max_size = 0
    list_result = [True, 'OK']
    for idx, item in enumerate(list):
        if idx == 0:
            set_results = set(item)
            if len(item) != len(set_results):
                list_result.append(False)
                list_result.append(dict_item[idx] + '에 중복된 입력값이 있습니다.')
                break
            max_size = len(item)
            continue
        else:
            if len(item) == max_size:
                pass
            else:
                list_result.append(False)
                list_result.append(dict_item[idx] + ' 의 값이 다른 config values 값과 상이합니다.')
                break

    return list_result



#Program Start~!

introText = """#####################################################-PACKET_BILLING-####################################################

 1. PACKET BILLING AUTO 결과 문서 출력 프로그램입니다.
 2. PC 바탕화면에 \"PB\" 폴더를 생성하세요
 3. PB 폴더 안에 \"INPUT_DATA\" 폴더를 생성 후 시험번호에서 산출된 \"CDR,UDR\" 파일을 넣어주세요.
 4. 반드시 결과 산출을 위해서는 시험번호 하나당 산출된 CDR,UDR 파일을 넣어주셔야 합니다.
 5. SET UP 문서의 \"SETUP\" Sheet에서 시험번호 정보와 시작시간/종료시간 을 정확히 넣어주세요.
 6. 결과 산출은 SETUP Sheet에 있는 결과 폴더명으로 \"바탕화면 / PB / \"에 생성 됩니다.

#########################################################################################################################"""+"\n\n"

home = expanduser("~")
nowTime = datetime.today().strftime("%Y%m%d_%H%M")
# Print Intro text
print(introText)

# Input and save user select each mode
inputFileName = input("\n=> SET UP Excel 파일명을(확장자 제외) 입력하여 주세요.(파일은 반드시 바탕화면 \"PB\" 폴더 안에 있어야 함) : ")
iNfileFlag = os.path.isfile(home+"\\Desktop\\PB\\"+inputFileName+".xlsx")

while not iNfileFlag:
    inputFileName = input("\n=> 입력하신 파일명이 경로에 존재하지 않습니다. 정확히 다시 입력하여 주세요. : ")
    iNfileFlag = os.path.isfile(home+"\\Desktop\\PB\\"+inputFileName+".xlsx")

inputCUFiles = input("\n=> PB > INPUT_DATA\" 폴더에 각 시험번호의 CDR,UDR 파일을 옮긴 후 \"y\"를 입력하시면 진행을 시작합니다.")
while inputCUFiles.lower() != "y":
    inputCUFiles = input("\n=> 입력이 올바르지 않습니다. \"y\"를 입력하세요.")

input_file = home+"\\Desktop\\PB\\"+inputFileName+".xlsx"
printText("\n SET UP 문서 경로 : {}".format(input_file))

# output_file = home+"\\Desktop\\VOC\\result_"+nowTime+".xlsx"
# print("==> 결과출력 문서 경로 : {}".format(output_file))

#Core Code

try:

    start_time = datetime.today().strftime("%Y-%m-%d %H:%M:%S")
    # Excel input Data read
    printText(inputFileName+" Started to read the SETUP File...")
    df_setup = pd.read_excel(input_file, sheet_name="SETUP", index_col=None, encoding='utf-8')

    # 시험명에 따른 시험명 리스트
    result_list = df_setup["시험명"].tolist()
    result_list = [str(x) for x in result_list if str(x) != 'nan']
    printText("SETUP 시험명 list 변환 완료")

    # 시험번호에 따른 시험번호 리스트
    tNumber_list = df_setup["시험번호"].tolist()
    tNumber_list = [str(x) for x in tNumber_list if str(x) != 'nan']
    tNumber_list = tNumberFormat(tNumber_list)
    printText("SETUP 시험번호 list 변환 완료")

    # 시험명에 따른 시작시간 리스트
    tStart_list = df_setup["시작시간"].tolist()
    tStart_list = [x for x in tStart_list if str(x) != 'nan']
    printText("SETUP 시작시간 list 변환 완료")

    # 시험명에 따른 종료시간 리스트
    tEnd_list = df_setup["종료시간"].tolist()
    tEnd_list = [x for x in tEnd_list if str(x) != 'nan']
    printText("SETUP 종료시간 list 변환 완료")

    # 시험명에 따른 Ipv6 Convert Port 리스트
    port_list = df_setup["IPv6 변환 Port"].tolist()
    port_list = [x for x in port_list if str(x) != 'nan']
    for j in range(len(port_list)):
        value = str(port_list[j])
        temp_list = value.split(",")
        port_list[j] = temp_list
    printText("SETUP Ipv6 port list 변환 완료")

    # 시험명에 따른 ip type 리스트
    ipType_list = df_setup["IPv6 변환 Type"].tolist()
    ipType_list = [x for x in ipType_list if str(x) != 'nan']
    printText("SETUP Ipv6 type list 변환 완료")

    list_total = [result_list, tNumber_list, tStart_list, tEnd_list, port_list, ipType_list]

    printText(inputFileName+" Finished to read the SETUP File...")

    # check config data
    printText(inputFileName+" Check the SETUP config values...")
    flag_check = checkConfigData(list_total)
    if not flag_check[0]:
        printText(flag_check[1])
        printText('Config 값을 다시 확인하시고 재실행 해주세요')
        sys.exit(-1)

    printText(inputFileName+" Start to analyze the INPUT_DATA File...")
    printText("시험번호 CDR/UDR 파일 존재 및 결과 도출 가능 여부 조회")

    inputFile_list = []
    cdrName_dic = {}
    udrName_dic = {}
    available_dic = {}
    disable_dic = {}
    netType_dic = {}
    tStart_dic = {}
    tEnd_dic = {}
    tNumber_dic = {}
    port_dic = {}
    ipType_dic = {}

    # UDR,CDR 파일 이름 읽어 들이기 작업~~~
    inputFile_list = os.listdir(home+"\\Desktop\\PB\\INPUT_DATA")
    inputFile_list.sort()
    # 시험번호에 따른 UDR, CDR 파일 분석 가능여부 확인 작업~~~
    for i in range(len(result_list)):

        available_list = []
        disable_list = []
        cdrFile_list = []
        udrFile_list = []
        tNumber_list = []
        tStart_list = []
        tEnd_list = []
        netType_list = []
        port_list = []
        ipType_list = []
        reduce_number = 0

        # #시험명에 따른 DataFrame set
        # df_temp = df_setup.loc[df_setup["시험구분"] == index_list[i],:]
        # df_temp =df_temp.reset_index(drop=True)

        # # ******************************시험명에 따른 망 리스트 삭제~~!!!******************************
        # netType_list = df_temp["망"].tolist()
        # netType_list = [x for x in netType_list if str(x) != 'nan']
        # ingString("시험명 ("+result_list[i]+") 망 list 변환 완료")
        # # ******************************시험명에 따른 망 리스트 삭제~~!!!******************************

        #시험번호 구분으로 분석가능 여부 조회
        printText("시험명 ("+result_list[i]+") 시험번호 분석가능 여부 학인 중")
        fCDR = False
        fUDR = False
        # pCDR = re.compile(".*"+result_list[i] +".*"+ tNumber_list[i]+".*CDR.*")
        # pUDR = re.compile(".*"+result_list[i] +".*"+ tNumber_list[i]+".*UDR.*")
        pCDR = re.compile(".*"+tNumber_list[i]+".*CDR.*")
        pUDR = re.compile(".*"+tNumber_list[i]+".*UDR.*")

        for item in inputFile_list:
            if pCDR.search(item):
                fCDR = True
            if pUDR.search(item):
                fUDR = True

        if fCDR and fUDR:
            available_list.append(tNumber_list[i])
            printText("("+result_list[i]+") 시험번호 "+tNumber_list[i]+" 분석 가능")

        else:
            reduce_number = reduce_number+1
            disable_list.append(tNumber_list[i])
                if fCDR == True and fUDR == False:
                    setPrintColor("yellow", "("+result_list[i]+") 시험번호 "+tNumber_list[j]+" UDR 파일이 \"INPUT_DATA\" 폴더에 존재하지 않습니다. 해당 번호는 분석을 하지 않습니다.")
                elif fCDR == False and fUDR == True:
                    setPrintColor("yellow", "("+result_list[i]+") 시험번호 "+tNumber_list[j]+" CDR 파일이 \"INPUT_DATA\" 폴더에 존재하지 않습니다. 해당 번호는 분석을 하지 않습니다.")
                else:
                    setPrintColor("yellow", "("+result_list[i]+") 시험번호 "+tNumber_list[j]+" CDR,UDR 파일이 \"INPUT_DATA\" 폴더에 존재하지 않습니다. 해당 번호는 분석을 하지 않습니다.")
                del netType_list[j-reduce_number]
                del tStart_list[j-reduce_number]
                del tEnd_list[j-reduce_number]

        #시험명에 따른 Dictionary Set
        available_dic[result_list[i]] = available_list
        cdrName_dic[result_list[i]] = cdrFile_list
        udrName_dic[result_list[i]] = udrFile_list
        disable_dic[result_list[i]] = disable_list
        tStart_dic[result_list[i]] = tStart_list
        netType_dic[result_list[i]] = netType_list
        tEnd_dic[result_list[i]] = tEnd_list
        port_dic[result_list[i]] = port_list
        ipType_dic[result_list[i]] = ipType_list

        # # 시험명 data dictionary 출력
        # setPrintColor("yellow",available_dic[result_list[i]])
        # setPrintColor("yellow",cdrName_dic[result_list[i]])
        # setPrintColor("yellow",udrName_dic[result_list[i]])
        # setPrintColor("yellow",disable_dic[result_list[i]])
        # setPrintColor("yellow",tStart_dic[result_list[i]])
        # setPrintColor("yellow",tEnd_dic[result_list[i]])
        # setPrintColor("yellow",netType_dic[result_list[i]])
        # setPrintColor("yellow",port_dic[result_list[i]])

    divString(inputFileName+" Finish to analyze the INPUT_DATA File")

    inputKeep = input("\n==> 계속 진행하시겠습니까?")
    while inputKeep.lower() != "y":
        inputKeep = input("\n==> 입력이 올바르지 않습니다. \"y\"를 입력하세요.")

    # 시험별 CDR UDR 파일 정리 작업~~~
    test_keys = list(cdrName_dic.keys())
    reCdrColumn_list = ["No", "ORG_FILE_NM", "data_type", "Served MSISDN", "Time of First Usage", "Time of Last Usage",\
                        "Data Volume FBC Uplink", "Data Volume FBC Downlink", "SUM", "Free Volume Uplink", "Free Volume Downlink",\
                        "Service Type",	"URL", "Destination IP", "Destination Port"]
    cdrColumn_list = ["No", "ORG_FILE_NM", "data_type", "Served MSISDN", "Time of First Usage", "Time of Last Usage",\
                        "Data Volume FBC Uplink", "Data Volume FBC Downlink", "Data Volume FBC SUM", "Free Volume Uplink", "Free Volume Downlink",\
                        "Service Type",	"URL", "Destination IP", "Destination Port"]
    udrColumn_list = ["No", "ORG_FILE_NM", "Data타입", "Served MN NAI", "PGW IP", "단말할당IP", "Access Point Name Network Identifier", "번호", "Start Time", "Stop Time",\
    	              "Duration", "Cause for Record Closing", "Served IMSI", "PDP/PDN Type", "Serving Node Address", "Serving node PLMN Identifier", "Service Key",\
                      "RAT Type", "User Location Information", "SKT-ZONE-INFO"]

    fotaIp_list = ["FOTA_Port","FOTA_Service_Type","FOTA_Ipv6","FOTA_Ipv4"]

    linkSum_list = ["pcdr_total_up","pcdr_total_down","pcdr_up/down_sum"]


    for item in test_keys:
        divString(item+" Start to Analyze Data")

        cdrFile_list   = cdrName_dic[item]
        udrFile_list   = udrName_dic[item]
        available_list = available_dic[item]
        tStart_list    = tStart_dic[item]
        tEnd_list      =   tEnd_dic[item]
        netType_list   = netType_dic[item]
        fotaPort_list = port_dic[item]
        fotaType_list = ipType_dic[item]
        output_path = home+"\\Desktop\\PB\\"+item+"("+nowTime+")"
        # executes mkdir method for result folder
        os.makedirs(output_path, exist_ok=True)

        #시험별 각 시험번호에 따라 작업~~~
        for i in range(len(udrFile_list)):
            # date_style = NamedStyle(name="datetime", number_format="YYYY-MM-DD")
            # DATA Frame
            df_udr = pd.read_excel(home+"\\Desktop\\PB\\INPUT_DATA\\"+udrFile_list[i], sheet_name="과금데이터조회", index_col=None, skiprows=2, encoding='utf-8')
            df_cdr = pd.read_excel(home+"\\Desktop\\PB\\INPUT_DATA\\"+cdrFile_list[i], sheet_name="과금데이터조회", index_col=None, skiprows=2, encoding='utf-8')

            #Calculate rows
            udrRows = df_udr.shape[0]
            cdrRows = df_cdr.shape[0]

            #Openpyxl set cdr/udr files
            wb_udr = openpyxl.load_workbook(home+"\\Desktop\\PB\\INPUT_DATA\\"+udrFile_list[i])
            wb_cdr = openpyxl.load_workbook(home+"\\Desktop\\PB\\INPUT_DATA\\"+cdrFile_list[i])
            udr_sheet = wb_udr.get_sheet_by_name('과금데이터조회')
            cdr_sheet = wb_cdr.get_sheet_by_name('과금데이터조회')

            string_a1 = udr_sheet["A1"].value
            string_a2 = cdr_sheet["A2"].value


            # # Auto Filter Set
            # udr_sheet.auto_filter.ref = "A3:T"+str(udrRows)
            # cdr_sheet.auto_filter.ref = "A3:N"+str(cdrRows)

            #Set udr start time ASC sort
            df_udr.sort_values(by=["Start Time"], axis=0, inplace=True)
            #After sort and reindex
            df_udr = df_udr.reset_index(drop=True)

            #Set cdr Time of First Usage ASC sort
            df_cdr.sort_values(by=["Time of First Usage"], axis=0, inplace=True)
            #After sort and reindex
            df_cdr = df_cdr.reset_index(drop=True)

            #Set string type of UDR's cell ==>  번호, start time, Stop time, Service IMSI, User Location Information, SKT-ZONE-INFO
            df_udr["번호"] = df_udr["번호"].astype(str)
            df_udr["Start Time"] = df_udr["Start Time"].astype(str)
            df_udr["Stop Time"] = df_udr["Stop Time"].astype(str)
            df_udr["Served IMSI"] = df_udr["Served IMSI"].astype(str)
            df_udr["User Location Information"] = df_udr["User Location Information"].astype(str)
            df_udr["SKT-ZONE-INFO"] = df_udr["SKT-ZONE-INFO"].astype(str)

            #Set string type of CDR's cell ==>  Served MSISDN, Time of First Usage, Time of Last Usage, Service IMSI, User Location Information, SKT-ZONE-INFO
            df_cdr["Served MSISDN"] = df_cdr["Served MSISDN"].astype(str)
            df_cdr["data_type"] = df_cdr["data_type"].astype(str)
            df_cdr["Time of First Usage"] = df_cdr["Time of First Usage"].astype(str)
            df_cdr["Time of Last Usage"] = df_cdr["Time of Last Usage"].astype(str)
            df_cdr["Data Volume FBC Uplink"] = df_cdr["Data Volume FBC Uplink"].astype(str)
            df_cdr["Data Volume FBC Downlink"] = df_cdr["Data Volume FBC Downlink"].astype(str)
            df_cdr["Free Volume Uplink"] = df_cdr["Free Volume Uplink"].astype(str)
            df_cdr["Free Volume Downlink"] = df_cdr["Free Volume Downlink"].astype(str)
            df_cdr["Destination Port"] = df_cdr["Destination Port"].astype(str)

            #시험번호 Start Time(dateType)
            startTime = tStart_list[i]
            #시험번호 Stop Time(dateType)
            endTime = tEnd_list[i]
            # udr start 시작시간 매칭 후 df_udrPaste 따로 저장
            df_udr["zFlag"] = ""
            for j in range(udrRows):
                tempTime  = str(df_udr.at[j,"Start Time"])
                hourTime = tempTime[8:10]
                minTime = tempTime[10:12]
                tempTime = time(int(hourTime), int(minTime))

                if tempTime >= startTime and tempTime <= endTime :
                    df_udr.at[j, "zFlag"] = 1
                else:
                    df_udr.at[j, "zFlag"] = 0


            #cdr의 Time of Last Usage가 종료시간보다 Over될 경우 해당 데이터 삭제
            # df = df.drop(df[(df.score < 50) & (df.score > 20)].index)
            # df = df[df["score"] > 50]
            df_cdr["zFlag"] = ""
            for j in range(cdrRows):
                tempTime  = str(df_cdr.at[j,"Time of Last Usage"])
                hourTime = tempTime[8:10]
                minTime = tempTime[10:12]
                tempTime = time(int(hourTime), int(minTime))

                if tempTime > endTime :
                    df_cdr.at[j, "zFlag"] = 1
                else:
                    df_cdr.at[j, "zFlag"] = 0

            #drop cdr row if zflag == 1
            df_cdr = df_cdr[df_cdr["zFlag"] == 0]
            df_cdr = df_cdr.reset_index(drop=True)
            df_cdr.drop('zFlag', axis=1, inplace=True)
            cdrRows = df_cdr.shape[0]

            #New UDR parsing for copy and paste to CDR File
            df_udrPaste = df_udr.loc[df_udr["zFlag"] == 1,:]
            df_udrPaste = df_udrPaste.reset_index(drop=True)
            # df_temp2 = df_keyword.loc[df_keyword["메모분류"] == item,:]

            #Delete 'zFlag' column df_udr and df_udrPaste
            df_udr.drop('zFlag', axis=1, inplace=True)
            df_udrPaste.drop('zFlag', axis=1, inplace=True)

            ingString("시험번호 "+available_list[i]+"번 UDR DATA_FRAME")
            print(df_udr)
            ingString("시험번호 "+available_list[i]+"번 CDR DATA_FRAME")
            print(df_cdr)

            # df_udrPaste = df_udrPaste.reset_index(drop=True)
            ingString("시험번호 "+available_list[i]+"번 UDR 시작시간 CONFIRM DATA_FRAME")
            print(df_udrPaste)

            #udrPaste copy and paste to CDR data frame
            row_list = []
            startTime_list = df_cdr["Time of First Usage"].tolist()
            udrPaste_start_list = df_udrPaste["Start Time"].tolist()
            udrPaste_end_list = df_udrPaste["Stop Time"].tolist()

            #compared cdr "Time of First Usage" and udrPaste_list Start Time
            for z in range(len(udrPaste_start_list)):

                for j in range(cdrRows):
                    first_time = df_cdr.at[j,"Time of First Usage"]
                    end_time = df_cdr.at[j,"Time of Last Usage"]
                    #Start time이 복수인 경우
                    if startTime_list.count(udrPaste_start_list[z]) > 1:
                        if first_time == udrPaste_start_list[z] and end_time == udrPaste_end_list[z-1]:
                            row_list.append(j+1)
                    #Start time이 단수인 경우
                    else:
                        if first_time == udrPaste_start_list[z]:
                            row_list.append(j)


            ingString("시험번호 "+available_list[i]+"번 MATCHING CDR ROW Numbers")
            print(row_list)

            # Add UDR Row to CDR top of the matching Time of First Usage cell
            # uplink plus downloadlink value and input in "Data Volume FBC SUM" cell
            df_result = df_cdr
            df_result["SUM"] = ""
            for j in range(df_result.shape[0]) :
                uplink = df_result.at[j, "Data Volume FBC Uplink"]
                downlink = df_result.at[j, "Data Volume FBC Downlink"]
                df_result.at[j,"SUM"] = int(uplink) + int(downlink)

            #Covert to String of  df_result "Data Volume FBC SUM" column
            df_result["SUM"] = df_result["SUM"].astype(str)
            df_result = df_result[reCdrColumn_list]

            #Set new column
            newColumn_list = ["UDR(Serving node PLMN Identifier)", "UDR(Service Key)", "UDR(RAT Type)", "UDR(User Location Information)","UDR(SKT-ZONE-INFO)"]
            setNewCol(newColumn_list, df_result)

            startTime_list = df_result["Time of First Usage"].tolist()
            refresh_list = []
            for j in range(len(row_list)) :
                df_temp = df_udrPaste.iloc[j]
                refresh_list.append(row_list[j]+j)
                df_result = insert_row(row_list[j]+j, df_result, df_temp)

            #Set cell type "Destination Port" and Delete "nan" value of "URL" column
            for j in range(df_result.shape[0]):
                port = df_result.at[j,"Destination Port"]
                dt = df_result.at[j,"data_type"]
                if "." in port:
                    pointIndex = port.find(".")
                    df_result.at[j,"Destination Port"] = port[:pointIndex]
                if "." in dt:
                    dtIndex = port.find(".")
                    df_result.at[j,"data_type"] = dt[:dtIndex]
                # url nan delete
                if df_result.at[j,"URL"] == "nan" or df_result.at[j,"URL"] == "":
                    df_result.at[j,"URL"] = ""

            #Set Cell type "Data Volume FBC Uplink" and "Data Volume FBC Downlink" and "SUM" convert to int
            # uplink = df_result["Data Volume FBC Uplink"]
            # uplink = pd.to_numeric(uplink, errors='ignore', downcast='integer')

            # downlink = df_result["Data Volume FBC Downlink"]
            # downlink = pd.to_numeric(downlink, errors='ignore', downcast='integer')

            # sumValue = df_result["SUM"]
            # sumValue = pd.to_numeric(sumValue, errors='ignore', downcast='integer')

            #Set uplink/downlink/sum cell sum value and generate dict
            sumUplink = 0
            sumDownlink = 0
            sumTotlink = 0
            for row in range(refresh_list[0],df_result.shape[0]):
                if row not in refresh_list:
                    sumUplink = sumUplink + int(df_result.at[row,"Data Volume FBC Uplink"])
                    # print("Up_link : "+df_result.at[row,"Data Volume FBC Uplink"])
                    sumDownlink = sumDownlink + int(df_result.at[row,"Data Volume FBC Downlink"])
                    # print("Down_link : "+df_result.at[row,"Data Volume FBC Downlink"])
                    sumTotlink = sumTotlink + int(df_result.at[row,"SUM"])
                    # print("Total_link : "+df_result.at[row,"SUM"])

            dict_sum = {"pcdr_total_up":[sumUplink],\
            "pcdr_total_down":[sumDownlink],\
            "pcdr_up/down_sum":[sumTotlink]
            }

            #Set service type "FRE" and "ETC" cell sum value and generate dict
            sumFre = 0
            sumEtc = 0
            for row in range(refresh_list[0],df_result.shape[0]):
                val_service = df_result.at[row,"Service Type"]
                val_totSum =  df_result.at[row,"SUM"]

                if val_service == "FRE":
                    sumFre = sumFre+int(val_totSum)
                elif val_service == "ETC":
                    sumEtc = sumEtc+int(val_totSum)


            dict_serviceSum = {"FRE":[sumFre],\
                "ETC":[sumEtc]
            }

            # Set value of IPV6 conver to integer Logic (Service Type이 "FRE"인 것 중에서 LG는 port==> 5005 Samsung은 ==> 80  )
            ipv6_list = []
            ipv4_list = []
            port_num_list = []
            service_list = []

            #ipv6 확인
            for j in range(refresh_list[0],df_result.shape[0]):
                val_serType = df_result.at[j,"Service Type"]
                val_Port = df_result.at[j,"Destination Port"]
                val_ip = df_result.at[j,"Destination IP"]
                fotaType_list[i] = fotaType_list[i].replace("Multi","")
                fotaType_list[i] = fotaType_list[i].replace("(","")
                fotaType_list[i] = fotaType_list[i].replace(")","")
                serviceSplit_list = fotaType_list[i].split("+")

                if val_serType in serviceSplit_list and  val_Port in fotaPort_list[i]:
                    ipv6_list.append(val_ip)
                    port_num_list.append(val_Port)
                    service_list.append(val_serType)

            #ipv6 Convert ipv4

            for port in ipv6_list:
                # num_flag = len(find_between_r(port,":",":"))
                split_list = port.split("::")
                val_target = split_list[1]
                val_target = val_target.replace(":","")

                address_part1 = ""
                address_part2 = ""
                address_part3 = ""
                address_part4 = ""

                if len(val_target) == 7:
                    address_part1 = val_target[0:1]
                    address_part1 = str(int(address_part1,16))
                    address_part2 = val_target[1:3]
                    address_part2 = str(int(address_part2,16))
                    address_part3 = val_target[3:5]
                    address_part3 = str(int(address_part3,16))
                    address_part4 = val_target[5:7]
                    address_part4 = str(int(address_part4,16))
                else:
                    address_part1 = val_target[0:2]
                    address_part1 = str(int(address_part1,16))
                    address_part2 = val_target[2:4]
                    address_part2 = str(int(address_part2,16))
                    address_part3 = val_target[4:6]
                    address_part3 = str(int(address_part3,16))
                    address_part4 = val_target[6:8]
                    address_part4 = str(int(address_part4,16))

                val_res = address_part1+"."+address_part2+"."+address_part3+"."+address_part4
                ipv4_list.append(val_res)


            dict_address = {"FOTA_Port":port_num_list,\
                "FOTA_Service_Type":service_list,\
                "FOTA_Ipv6":ipv6_list,\
                "FOTA_Ipv4":ipv4_list
            }
            #generate DataFrame
            df_ipAddress = pd.DataFrame(dict_address)
            df_sum = pd.DataFrame(dict_sum)
            df_serviceSum = pd.DataFrame(dict_serviceSum)

            #column sort
            df_ipAddress = df_ipAddress[fotaIp_list]
            df_sum = df_sum[linkSum_list]
            df_serviceSum = df_serviceSum[["FRE","ETC"]]


            # {available_list}_CDR+UDR_{netType_list}
            writer = pd.ExcelWriter(output_path+"\\"+available_list[i]+"_CDR+UDR_"+netType_list[i]+".xlsx")
            df_result.to_excel(writer, sheet_name="과금데이터조회", index=False, startrow=2)
            df_ipAddress.to_excel(writer, sheet_name="Summary", index=False)
            df_serviceSum.to_excel(writer, sheet_name="Summary", index=False, startcol=4)
            df_sum.to_excel(writer, sheet_name="Summary", index=False, startcol=6)

            writer.save()

#############################################################################Openpyxl#######################################################################################
            ingString(available_list[i]+" Start to set Excel Options")
            #Openpyxl set cdr/udr files
            wb_res = openpyxl.load_workbook(output_path+"\\"+available_list[i]+"_CDR+UDR_"+netType_list[i]+".xlsx")
            res_sheet = wb_res.get_sheet_by_name('과금데이터조회')
            sum_sheet = wb_res.get_sheet_by_name('Summary')
            res_sheet['A1'] = string_a1
            res_sheet['A2'] = string_a2
            resRows = df_result.shape[0]+3
            sumRows = df_ipAddress.shape[0]+1

            res_cell_list = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N',\
            'O','P','Q','R','S','T']
            sum_cell_list = ['A','B','C','D','E','F','G','H','I']

            res_width_list = [6.88, 11.25, 8, 15.63, 15.63, 15.63, 13.25, 13.25, 13.25, 15.38, 15.38, 10.7, 14.13, 14.25, 14.25, 20.88\
            ,15.13, 15.13, 20.88, 15.13]

            sum_width_list = [18, 18, 14, 18, 17.38, 17.38, 19, 19, 19]

            box = Border(left=Side(border_style="thin",\
                        color='FF000000'),\
                        right=Side(border_style="thin",\
                        color='FF000000'),\
                        top=Side(border_style="thin",\
                        color='FF000000'),\
                        bottom=Side(border_style="thin",\
                        color='FF000000'),\
                        diagonal=Side(border_style="thin",\
                        color='FF000000'),\
                        diagonal_direction=0,\
                        outline=Side(border_style="thin",\
                        color='FF000000'),\
                        vertical=Side(border_style="thin",\
                        color='FF000000'),\
                        horizontal=Side(border_style="thin",\
                        color='FF000000')\
                    )

            # set width value for header
            for j in range(len(res_width_list)):
                res_sheet.column_dimensions[res_cell_list[j]].width = res_width_list[j]
            for j in range(len(sum_width_list)):
                sum_sheet.column_dimensions[sum_cell_list[j]].width = sum_width_list[j]

            # Each Sheet set auto filter column
            res_sheet.auto_filter.ref = "A3:T"+str(resRows)
            sum_sheet.auto_filter.ref = "A1:I"+str(sumRows)

            # Set A1,A2 options
            res_sheet['A1'].font = Font(name='맑은 고딕', size=13, bold=True, color='000000')
            res_sheet['A2'].font = Font(name='맑은 고딕', size=13, bold=True, color='000000')

            # set hole Excel option for res_sheet
            for mCell in res_sheet["A3:T"+str(resRows)]:
                    for cell in mCell:
                        if cell.row == 3:
                            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
                            cell.font = Font(name='맑은 고딕', size=10, bold=True, color='ffffff')
                        else:
                            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
                            cell.font = Font(size=10)

            # set hole Excel option for sum_sheet
            for mCell in sum_sheet["A1:I"+str(sumRows)]:
                    for cell in mCell:
                        if cell.row == 1:
                            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
                            cell.font = Font(name='맑은 고딕', size=10, bold=True, color='ffffff')
                        else:
                            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
                            cell.font = Font(size=10)

            #set Cell Back ground and Border for res_sheet
            target_list = []
            for num in refresh_list:
                value = num+4
                target_list.append(value)

            for col in res_cell_list:
                for row in range(3,resRows+1):
                    cell = res_sheet[col+str(row)]
                    if row is 3:
                        cell.fill = cell.fill = PatternFill(patternType= 'solid', fill_type= 'solid', fgColor= Color('808080'))
                    if row is 3 and (col == "P" or col == "Q" or col == "R" or col == "S" or col == "T"):
                        cell.fill = cell.fill = PatternFill(patternType= 'solid', fill_type= 'solid', fgColor= Color('0070c0'))
                    if row in target_list:
                        cell.fill = cell.fill = PatternFill(patternType= 'solid', fill_type= 'solid', fgColor= Color('b4c6e7'))
                    cell.border = box

            #set Cell Back ground and Border for sum_sheet
            fota_col = ['A','B','C','D']
            service_col = ['E','F']
            sum_cols = ['G','H','I']
            for col in sum_cell_list:
                for row in range(1,sumRows+1):
                    cell = sum_sheet[col+str(row)]

                    #fota_col Header set Style
                    if col in fota_col and row is 1:
                        cell.fill = cell.fill = PatternFill(patternType= 'solid', fill_type= 'solid', fgColor= Color('95b3d7'))
                        cell.border = box

                    #service_col Header and values set Style
                    if col in service_col and row < 3:
                        if row is 1 and col == "F":
                            cell.fill = cell.fill = PatternFill(patternType= 'solid', fill_type= 'solid', fgColor= Color('fcd5b4'))
                            cell.font = Font(color="ff0000", size=10, bold=True)
                            cell.border = box
                        elif row is 1 and col == "E":
                            cell.fill = cell.fill = PatternFill(patternType= 'solid', fill_type= 'solid', fgColor= Color('fcd5b4'))
                            cell.font = Font(color="0000ff", size=10, bold=True)
                            cell.border = box
                        elif row is 2 and col == "F":
                            cell.font = Font(color="ff0000", size=10, bold=False)
                            cell.alignment = Alignment(horizontal="right")
                            cell.number_format = "#,##0"
                            cell.border = box
                        elif row is 2 and col == "E":
                            cell.font = Font(color="0000ff", size=10, bold=False)
                            cell.alignment = Alignment(horizontal="right")
                            cell.number_format = "#,##0"
                            cell.border = box

                    #sum_col Header and values set Style
                    if col in sum_cols and row < 3:
                        if row is 1:
                            cell.fill = cell.fill = PatternFill(patternType= 'solid', fill_type= 'solid', fgColor= Color('d8e4bc'))
                            cell.font = Font(color="000001", size=9, bold=True)
                        if row is 2:
                            cell.alignment = Alignment(horizontal="right")
                            cell.font = Font(color="000001", size=10, bold=False)
                            cell.number_format = "#,##0"
                        cell.border = box

                    #fota_col values set Style
                    if (col not in sum_cols and col not in service_col) and row > 1:
                        cell.border = box

            #set font color "Red" and "Blue" on "Service Type" if ETC and FRE for res_sheet
            startRow = refresh_list[0]+4
            for row in range(startRow, resRows+1):
                cell = res_sheet["L"+str(row)]
                if cell.value == "ETC":
                    for col in res_cell_list:
                        cell2 = res_sheet[col+str(row)]
                        cell2.font = Font(name='맑은 고딕', size=10, bold=False, color='ff0000')
                if cell.value == "FRE":
                    for col in res_cell_list:
                        cell2 = res_sheet[col+str(row)]
                        cell2.font = Font(name='맑은 고딕', size=10, bold=False, color='0000ff')


            wb_res.save(output_path+"\\"+available_list[i]+"_CDR+UDR_"+netType_list[i]+".xlsx")
            ingString(available_list[i]+" Finish to set Excel Options")

        divString(item+" Finish to Analyze Data")
        print("\nAll job is completed.")
        input("\nPress any key to exit...")

except:
    print("Unexpected error:", sys.exc_info())
    input("\nPress any key to exit...")
