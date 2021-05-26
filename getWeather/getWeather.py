# -*- coding:utf-8 -*-

import os
import sys
import pymysql
import openpyxl
import urllib
import requests
import math
import json
import xmltodict
# import xml.etree.ElementTree as eleTree

# import threading
# import pandas as pd
# import tkfilebrowser
# from tqdm import tqdm

from openpyxl.styles import Alignment, Font, NamedStyle, PatternFill
from openpyxl import formatting, styles, Workbook
from openpyxl.styles.borders import Border, Side
from time import sleep
from os.path import expanduser
from datetime import datetime
from datetime import timedelta
from apscheduler.schedulers.background import BackgroundScheduler


# 요청 형식
# 일기예보
# http://apis.data.go.kr/1360000/VilageFcstInfoService/getVilageFcst?serviceKey=96uh0W7mDdbpm2cUHdS12hg6HGxDKjQ4GU7DDCSFS8eKgocNkWdTdxnPXkZwwj7kOevMYo2W37VF0xqocdj3TA%3D%3D&numOfRows=300&pageNo=1&base_date=20210217&base_time=0000&nx=55&ny=108
# 초단기 날씨실황
# http://apis.data.go.kr/1360000/VilageFcstInfoService/getUltraSrtNcst?serviceKey=96uh0W7mDdbpm2cUHdS12hg6HGxDKjQ4GU7DDCSFS8eKgocNkWdTdxnPXkZwwj7kOevMYo2W37VF0xqocdj3TA%3D%3D&numOfRows=3 00&pageNo=1&base_date=20210217&base_time=0000&nx=55&ny=108
# 초단기 일기예보
# http://apis.data.go.kr/1360000/VilageFcstInfoService/getUltraSrtFcst?serviceKey=96uh0W7mDdbpm2cUHdS12hg6HGxDKjQ4GU7DDCSFS8eKgocNkWdTdxnPXkZwwj7kOevMYo2W37VF0xqocdj3TA%3D%3D&numOfRows=3 00&pageNo=1&base_date=20210217&base_time=1141&nx=55&ny=108


class GetWeather():
    # 클레스 초기화
    def __init__(self):

        super().__init__()
        self.current_path = os.getcwd()
        self.home = expanduser("~")
        self.config_path = self.current_path+"\\GET_WEATHER_CONFIG.xlsx"
        self.conn = None
        self.cur = None
        self.local_names = []
        self.local_code = []
        self.city_names = []
        self.weather_schedules = []
        self.forecast_schedules = []
        self.shortcast_schedules = []
        self.riseset_schedules = []
        self.weather_address = None
        self.forecast_address = None
        self.shortcast_address = None
        self.riseset_address = None
        self.service_key = None
        self.service_key_riseset = None
        self.db_id = 'root'
        self.db_pw = 'navy1063'
        self.db_name = 'beanfarm'
        self.table_weather_name = 'weather'
        self.table_forecast_name = 'forecast'
        self.hold_flag = False

    # 시간 정보와 함께 Console에 Print 하는 함수
    def set_print(self, text):

        current = datetime.today().strftime("%Y-%m-%d %H:%M:%S")
        print("{}:\n{}\n".format(current, text))

    # 현재 시간 구하여 3가지 타입으로 list return
    def get_currentTime(self):

        nowTime = datetime.now()
        nowtime_str = nowTime.strftime('%Y-%m-%d %H:%M:%S')
        nowtime_str_2 = nowTime.strftime('%Y-%m-%d %H %M %S')
        list_now = [nowTime.strftime('%Y%m%d'), nowTime.strftime('%H%M')]
        return [nowTime, nowtime_str, nowtime_str_2, list_now]

    # Base Date 및 Base time fomatter
    def change_format_bases(self, base_date, base_time):

        reform_base_date = datetime.strptime(base_date, '%Y%m%d')
        return_base_date = reform_base_date.strftime('%Y-%m-%d')
        reform_base_time = datetime.strptime(base_time, '%H%M')
        return_base_time = reform_base_time.strftime('%H:%M')

        return return_base_date+' '+return_base_time

    # 구분자 포함을 확인한 후 배열로 리턴
    def set_values(self, value):

        check_value = str(value).strip()
        if check_value == 'nan' or check_value == '':
            return []
        else:
            if ',' in check_value:
                list_data = check_value.split(',')
                list_data = [x.strip() for x in list_data]
                return list_data
            else:
                return [check_value.strip()]

    # 최신 지역정보 DB로 부터 가져오기
    def get_localData(self):

        self.conncet_db(message='지역정보')
        try:
            # mysql 지역정보 가져오기 진행
            self.set_print('DB 지역/도서 정보 서버로부터 수신 진행')
            sql = "SELECT * FROM locals"
            self.cur.execute(sql)
            rows = self.cur.fetchall()
            data_num = len(rows)
            self.set_print('DB 지역/도서 수신 정보 {} 개 데이터 parsing 진행'.format(data_num))
            # Each row local data parsing
            for row in rows:
                use_flag = row['FLAG']
                if use_flag != 1:
                    continue
                state1 = row['STATE1']
                state2 = row['STATE2']
                state3 = row['STATE3']
                city_name = row['LOCAL']
                self.local_names.append([state1, state2, state3])
                self.city_names.append(city_name)

            self.set_print('DB 지역/도서 정보 서버로부터 수신 완료')

        except:
            self.set_print('Error: {}. {}, line: {}'.format(sys.exc_info()[0], sys.exc_info()[1], sys.exc_info()[2].tb_lineno))

        finally:
            self.set_print('지역정보 DB 연결 해제')
            self.cur.close()
            self.conn.close()

    # riseset 5일치 Data 날짜 list 얻기
    def get_riseset_list_date(self, recent_date, include=False, count=5):

        list_return_date = []
        recent_to_dateFormat = datetime.strptime(recent_date, '%Y%m%d')

        if include:
            for i in range(0, count):
                item_date = recent_to_dateFormat + timedelta(days=i)
                item_date = item_date.strftime('%Y%m%d')
                list_return_date.append(item_date)
        else:
            for i in range(1, count+1):
                item_date = recent_to_dateFormat + timedelta(days=i)
                item_date = item_date.strftime('%Y%m%d')
                list_return_date.append(item_date)

        return list_return_date

    # 최신 riseset DB로 부터 가져오기
    def get_riseset_recent_date(self, city):

        recent_date = None
        try:
            # mysql riseset 가져오기 진행
            self.set_print('DB riseset 정보 서버로부터 수신 진행')
            sql = "SELECT LOCDATE FROM riseset WHERE LOCATION=\""+city+"\" ORDER BY LOCDATE DESC , LOCATION ASC LIMIT 1"
            self.cur.execute(sql)
            rows = self.cur.fetchall()
            data_num = len(rows)
            # Each row local data parsing
            if data_num != 0:
                recent_date = rows[0]['LOCDATE']
                self.set_print('DB riseset 최근일자 {} 수신 완료'.format(recent_date))
            else:
                self.set_print('DB riseset 최근일자가 없음. 금일 base date로 설정')

        except:
            self.set_print('Error: {}. {}, line: {}'.format(sys.exc_info()[0], sys.exc_info()[1], sys.exc_info()[2].tb_lineno))

        finally:
            return recent_date

    # 모든 sheet 이름 추출
    def check_sheetNames(self, list_data):

        return_status = [True, 'ok']
        if '지역코드표' not in list_data:
            return_status[0] = False
            return_status[1] = 'config 파일에 "지역코드표" Sheet가 존재하지 않습니다.'
            return return_status
        if '지역설정' not in list_data:
            return_status[0] = False
            return_status[1] = 'config 파일에 "지역설정" Sheet가 존재하지 않습니다.'
            return return_status
        if '일기예보시간설정' not in list_data:
            return_status[0] = False
            return_status[1] = 'config 파일에 "일기예보시간설정" Sheet가 존재하지 않습니다.'
            return return_status
        if '단기예보시간설정' not in list_data:
            return_status[0] = False
            return_status[1] = 'config 파일에 "단기예보시간설정" Sheet가 존재하지 않습니다.'
            return return_status
        if '단기실황시간설정' not in list_data:
            return_status[0] = False
            return_status[1] = 'config 파일에 "단기실황시간설정" Sheet가 존재하지 않습니다.'
            return return_status
        if '일출시간설정' not in list_data:
            return_status[0] = False
            return_status[1] = 'config 파일에 "일출시간설정" Sheet가 존재하지 않습니다.'
            return return_status
        if '주소정보(일기)' not in list_data:
            return_status[0] = False
            return_status[1] = 'config 파일에 "주소정보(일기)" Sheet가 존재하지 않습니다.'
            return return_status
        if '주소정보(천문)' not in list_data:
            return_status[0] = False
            return_status[1] = 'config 파일에 "주소정보(천문)" Sheet가 존재하지 않습니다.'
            return return_status
        if 'DB정보' not in list_data:
            return_status[0] = False
            return_status[1] = 'config 파일에 "DB정보" Sheet가 존재하지 않습니다.'
        return return_status

    # 모든 sheet 데이터 유무 조회
    def check_sheetData(self, list_names, list_sheets):

        return_status = [True, 'ok']
        for idx, sheet in enumerate(list_sheets):
            max_rows = sheet.max_row
            if max_rows < 2:
                return_status[0] = False
                return_status[1] = '"{}" sheet의 데이터가 존재하지 않습니다.'.format(list_names[idx])
                break
        return return_status

    # Config Excel 파일 read 함수
    def get_configFile(self):

        self.set_print('설정 데이터 Read...')
        wb = openpyxl.load_workbook(self.config_path, data_only=True)
        list_sheet_names = wb.sheetnames
        # check mandatory sheet tab names
        flag_sheet = self.check_sheetNames(list_sheet_names)
        if not flag_sheet[0]:
            self.set_print(flag_sheet[1])
            sys.exit(-1)

        sheet_local_code = wb['지역코드표']
        sheet_local_set = wb['지역설정']
        sheet_forecast_set = wb['일기예보시간설정']
        sheet_shortcast_set = wb['단기예보시간설정']
        sheet_weather_set = wb['단기실황시간설정']
        sheet_riseset_set = wb['일출시간설정']
        sheet_address_set = wb['주소정보(일기)']
        sheet_address_riseset_set = wb['주소정보(천문)']
        sheet_db_set = wb['DB정보']

        # check sheet data exists
        flag_data = self.check_sheetData(list_sheet_names,
                                         [sheet_local_code, sheet_local_set, sheet_forecast_set, sheet_shortcast_set,
                                          sheet_weather_set, sheet_riseset_set, sheet_address_set,
                                          sheet_address_riseset_set])
        if not flag_data[0]:
            self.set_print(flag_datg[1])
            sys.exit(-1)

        # 지역정보 서버로 부터 획득 함수 호출
        self.get_localData()

        # get code nx/ny code total
        max_rows = sheet_local_code.max_row
        for item in self.local_names:

            temp_local_code = ['0', '0']
            # list local set에 있는 지명을 지역코드 표에서 대조하여 찾아 nx와 ny를 local_code에 입력
            for idx in range(2, max_rows + 1):

                state1 = sheet_local_code['C' + str(idx)].value
                if not state1:
                    state1 = None
                state2 = sheet_local_code['D' + str(idx)].value
                if not state2:
                    state2 = None
                state3 = sheet_local_code['E' + str(idx)].value
                if not state3:
                    state3 = None


                # 만약 state1/2/3 지역이 모두 일치할 경우 nx/ny 입력
                if item[0] == state1 and item[1] == state2 and item[2] == state3:
                    nx = str(sheet_local_code['F' + str(idx)].value)
                    ny = str(sheet_local_code['G' + str(idx)].value)
                    temp_local_code[0] = nx
                    temp_local_code[1] = ny
                    break

            # 만약 지역명이 틀려서 코드를 얻지 못하는 경우
            if temp_local_code[0] == '0' or temp_local_code[1] == '0':
                self.set_print('지역설정 {} 지역명이 잘못되어 있습니다.'.format(item))
                sys.exit(-1)

            # nx/ny 좌표 값 local_code에 입력
            self.local_code.append(temp_local_code)

        # 일기예보시간 설정
        schedule_type = sheet_forecast_set['A2'].value.strip()
        schedule_time = self.set_values(sheet_forecast_set['B2'].value)
        schedule_interval = sheet_forecast_set['C2'].value
        self.forecast_schedules = [schedule_type, schedule_time, schedule_interval]

        # 단기예보시간 설정
        schedule_type = sheet_shortcast_set['A2'].value.strip()
        schedule_time = self.set_values(sheet_shortcast_set['B2'].value)
        schedule_interval = sheet_shortcast_set['C2'].value
        self.shortcast_schedules = [schedule_type, schedule_time, schedule_interval]

        # 단기실황시간설정
        schedule_type = sheet_weather_set['A2'].value
        schedule_time = self.set_values(sheet_weather_set['B2'].value)
        schedule_interval = sheet_weather_set['C2'].value
        self.weather_schedules = [schedule_type, schedule_time, schedule_interval]

        # 일출시간설정
        schedule_type = sheet_riseset_set['A2'].value
        schedule_time = self.set_values(sheet_riseset_set['B2'].value)
        schedule_interval = sheet_riseset_set['C2'].value
        self.riseset_schedules = [schedule_type, schedule_time, schedule_interval]

        # 주소정보(일기) 및 서비스 key
        self.forecast_address = sheet_address_set['B2'].value
        self.shortcast_address = sheet_address_set['B4'].value
        self.weather_address = sheet_address_set['B3'].value
        self.service_key = sheet_address_set['C2'].value

        # 주소정보(천문) 및 서비스 key
        self.riseset_address = sheet_address_riseset_set['B2'].value
        self.service_key_riseset = sheet_address_riseset_set['C2'].value


        # None 값이 있는지 확인
        if not self.forecast_address:
            self.set_print('주소정보(일기) sheet에 일기예보 주소 값이 없습니다.')
            sys.exit(-1)
        if not self.weather_address:
            self.set_print('주소정보(일기) sheet에 초단기 주소 값이 없습니다.')
            sys.exit(-1)
        if not self.shortcast_address:
            self.set_print('주소정보(일기) sheet에 단기예보 주소 값이 없습니다.')
            sys.exit(-1)
        if not self.service_key:
            self.set_print('주소정보(일기) sheet에 service key 값이 없습니다.')
            sys.exit(-1)
        if not self.riseset_address:
            self.set_print('주소정보(천문) sheet에 service key 값이 없습니다.')
            sys.exit(-1)
        if not self.service_key_riseset:
            self.set_print('주소정보(천문) sheet에 service key 값이 없습니다.')
            sys.exit(-1)

        # DB 정보 설정
        account_id = sheet_db_set['B2']
        account_pw = sheet_db_set['B3']
        db_name = sheet_db_set['B4']
        t1 = sheet_db_set['B5']
        t2 = sheet_db_set['B6']

        if not account_id:
            self.db_id = account_id
        if not account_pw:
            self.db_pw = account_pw
        if not db_name:
            self.db_name = db_name
        if not t1:
            self.table_weather_name = t1
        if not account_id:
            self.table_forecast_name = t2
        self.set_print('설정 데이터 저장 완료...')

    # mysql db generate connection
    def conncet_db(self, message=''):

        try:
            self.set_print('{} DB Connection 작업 진행'.format(message))
            self.conn = pymysql.connect(host='localhost', user=self.db_id, password=self.db_pw,
                                        db=self.db_name, charset='utf8', autocommit=True)
            self.cur = self.conn.cursor(pymysql.cursors.DictCursor)
            self.set_print('{} DB 연동 Connection 작업 완료'.format(message))

        except:
            self.set_print('Error Occurred : {}. {}, line: {}'.format(sys.exc_info()[0], sys.exc_info()[1],
                                                                      sys.exc_info()[2].tb_lineno))

    # RN1 or R06 data format change
    def change_format_rn1(self, value):

        return_data = None
        measure_value = float(value)
        dict_test = {(0.1, 1.0): '1mm 미만', (1.0, 5.0): '1~4mm', (5.0, 10.0): '5~9mm', (10.0, 20.0): '10~19mm',
                     (20.0, 40.0): '20~39mm', (40.0, 70.0): '40~69mm'}

        if measure_value < 0.1:
            return_data = '강수 없음'
        elif measure_value >= 70.0:
            return_data = '70mm 이상'
        else:
            for key, value in dict_test.items():
                tuple_range = key
                if (tuple_range[0] <= measure_value) and (tuple_range[1] > measure_value):
                    return_data = value
        return return_data

    # S06 data format change
    def change_format_s06(self, value):

        return_data = None
        measure_value = float(value)
        dict_test = {(0.1, 1.0): '1cm 미만', (1.0, 5.0): '1~4cm', (5.0, 10.0): '5~9cm', (10.0, 20.0): '10~19cm'}

        if measure_value < 0.1:
            return_data = '적설 없음'
        elif measure_value >= 20.0:
            return_data = '20cm 이상'
        else:
            for key, value in dict_test.items():
                tuple_range = key
                if (tuple_range[0] <= measure_value) and (tuple_range[1] > measure_value):
                    return_data = value
        return return_data

    # PTY data format change
    def change_format_pty(self, value):

        dict_criterion = {'0': '없음', '1': '비', '2': '비/눈', '3': '눈', '4': '소나기', '5': '빗방울',
                          '6': '빗방울/눈날림', '7': '눈날림'}
        change_value = dict_criterion[value]
        return change_value

    # SKY data format change
    def change_format_sky(self, value):

        return_data = '맑음'
        if value == '2':
            return_data = '구름조금'
        elif value == '3':
            return_data = '구름많음'
        elif value == '4':
            return_data = '흐림'

        return return_data

    # LGT data format change
    def change_format_lgt(self, value):

        return_data = '없음'
        if value == '1':
            return_data = '있음'
        elif value == '2':
            return_data = '보통'
        elif value == '3':
            return_data = '높음'

        return return_data

    # UUU data format change
    def change_format_uuu(self, value):
        direction = '동'
        speed = str(value)
        if '-' in value:
            direction = '서'
            speed = str(value).replace('-', '')

        return direction+' '+speed

    # UUU data format change
    def change_format_vvv(self, value):
        direction = '북'
        speed = str(value)
        if '-' in value:
            direction = '남'
            speed = str(value).replace('-', '')

        return direction + ' ' + speed

    # vec data format change
    def change_format_vec(self, value):

        dict_vec = {0: 'N', 1: 'NNW', 2: 'NE', 3: 'ENE', 4: 'E', 5: 'ESE', 6: 'SE', 7: 'SSE', 8: 'S',
                    9: 'SSW', 10: 'SW', 11: 'WSW', 12: 'W', 13: 'WNW', 14: 'NW', 15: 'NNW', 16: 'N'}
        vector_deg = int(value)
        vector = float((vector_deg + 22.5*0.5)/22.5)
        vector = math.trunc(vector)
        return_data = dict_vec[vector]

        return return_data

    # 초단기 날씨 현황 처리 fucntion
    def forecast_handler(self):

        # 배열 1st ==> 응답코드+응답메세지
        # 배열 2nd ==> items dict가 모여있는 list(없는 경우 비어 있는 list 값)
        # 배열 3rd ==> items 데이터 개수
        self.hold_flag = True
        self.conncet_db('일기예보 정보')
        try:
            header = {'Content-Type': 'application/json; charset=utf-8'}
            current_time = self.get_currentTime()
            base_date_request = current_time[3][0]
            base_time_request = current_time[3][1]
            list_col = ['UPLOAD_TIME', 'POP', 'PTY', 'R06', 'REH', 'S06', 'SKY', 'T3H', 'TMN', 'TMX', 'UUU',
                        'VVV', 'WAV', 'VEC', 'WSD', 'NX', 'NY', 'STATE1', 'STATE2', 'STATE3', 'LOCATION',
                        'BASE_DATETIME', 'FCST_DATETIME']
            list_measure = ['POP', 'PTY', 'R06', 'REH', 'S06', 'SKY', 'T3H', 'TMN', 'TMX', 'UUU', 'VVV', 'WAV',
                            'VEC', 'WSD']

            for idx, code in enumerate(self.local_code):
                # 각 변수 초기화
                response_data = []
                state1 = self.local_names[idx][0]
                state2 = self.local_names[idx][1]
                state3 = self.local_names[idx][2]
                location = self.city_names[idx]

                list_items = []
                list_fcst_tuples = []

                sql_insert = "REPLACE INTO forecast("+', '.join(list_col)+") \
                VALUES(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                ########################################__실행해야 하는 지역코드 만큼 get api 실행 영역__########################################
                dict_parameter = 'serviceKey='+self.service_key+'&base_date='+base_date_request+'&base_time='+base_time_request+\
                                 '&nx='+code[0]+'&ny='+code[1]+'&numOfRows=300&pageNo=1&dataType=JSON'

                response = requests.get(self.forecast_address, headers=header, params=dict_parameter)
                self.set_print('{} {} {} 지역일기예보 API 요청\n[REQUEST]\n{}'.format(state1, state2, state3,
                                                                              response.request.url))
                self.set_print('{} {} {} 지역일기예보 API 응답\n[RESPONSE]\nHeader:{}\nbody:{}'.format(state1, state2, state3,
                                                                               response.headers, response.text))
                # status_code 정상인 경우
                if response.status_code == 200:

                    response_json = response.json()
                    res_code = response_json['response']['header']['resultCode']
                    res_message = response_json['response']['header']['resultMsg']
                    # 정상 처리 됨
                    if res_code == '00':
                        items = response_json['response']['body']['items']['item']
                        data_size = response_json['response']['body']['totalCount']
                        response_data.append(res_code+' '+res_message)
                        response_data.append(items)
                        response_data.append(data_size)
                    # response code 에러 발생 비정상 처리
                    else:
                        raise Exception(
                            'REQUEST {} 에서 Error 발생 \n response code: {}\n response message: {}'.format
                            (response.request.url, res_code, res_message))
                # status_code 비정상의 경우
                else:
                    response.raise_for_status()
                ############################################__items parsing 작업 실행 영역__############################################
                return_base = self.change_format_bases(response_data[1][0]['baseDate'], response_data[1][0]['baseTime'])

                # list unique fcst_date and fcst_time data
                for item in response_data[1]:

                    fcst_data = (item['fcstDate'], item['fcstTime'])
                    if fcst_data not in list_fcst_tuples:
                        list_fcst_tuples.append(fcst_data)

                # mysql dict에 response data 넣어 주기
                for item in list_fcst_tuples:

                    creterion_fcst_date = item[0]
                    creterion_fcst_time = item[1]

                    return_fcst = self.change_format_bases(creterion_fcst_date, creterion_fcst_time)
                    mysql_items = {'UPLOAD_TIME': current_time[1], 'NX': code[0], 'NY': code[1], 'STATE1': state1,
                                   'STATE2': state2, 'STATE3': state3, 'BASE_DATETIME': return_base,
                                   'FCST_DATETIME': return_fcst, 'LOCATION': location}

                    for item2 in response_data[1]:

                        if item2['fcstDate'] == creterion_fcst_date and item2['fcstTime'] == creterion_fcst_time:
                            category = item2['category']
                            value = item2['fcstValue']
                            mysql_items[category] = str(value)

                    # data format change if need and check no data check
                    mysql_items['SKY'] = self.change_format_sky(mysql_items['SKY'])
                    mysql_items['PTY'] = self.change_format_pty(mysql_items['PTY'])
                    mysql_items['VEC'] = self.change_format_vec(mysql_items['VEC'])
                    mysql_items['VVV'] = self.change_format_vvv(mysql_items['VVV'])
                    mysql_items['UUU'] = self.change_format_uuu(mysql_items['UUU'])

                    # if some data is none put in text as 'No data'
                    list_keys = list(mysql_items.keys())
                    if 'R06' not in list_keys:
                        mysql_items['R06'] = 'No Data'
                    else:
                        mysql_items['R06'] = self.change_format_rn1(mysql_items['R06'])

                    if 'S06' not in list_keys:
                        mysql_items['S06'] = 'No Data'
                    else:
                        mysql_items['S06'] = self.change_format_s06(mysql_items['S06'])

                    if 'TMN' not in list_keys:
                        mysql_items['TMN'] = 'No Data'
                    if 'TMX' not in list_keys:
                        mysql_items['TMX'] = 'No Data'
                    if 'WAV' not in list_keys:
                        mysql_items['WAV'] = 'No Data'

                    ############################################__MySQL DB에 Upload 데이터 실행 영역__###########################################
                    # mysql tuple mysql insert 순서로 데이터 생성
                    list_temp_mysql = []
                    for col in list_col:
                        list_temp_mysql.append(mysql_items[col])
                    mysql_tuple = tuple(list_temp_mysql)
                    list_items.append(mysql_tuple)

                # mysql 업로드 진행
                self.set_print('{} {} {} DB 지역일기예보 정보 {}건 업로드 진행'.format(state1, state2, state3, len(list_items)))
                for tuple_data in list_items:
                    self.cur.execute(sql_insert, tuple_data)
                self.conn.commit()
                self.set_print('{} {} {} DB 지역일기예보 정보 업로드 완료'.format(state1, state2, state3))

        except:
            self.set_print('Error: {}. {}, line: {}'.format(sys.exc_info()[0], sys.exc_info()[1], sys.exc_info()[2].tb_lineno))

        finally:
            self.set_print('일기예보 정보 DB 연결 해제')
            self.cur.close()
            self.conn.close()
            self.hold_flag = False

    # 초단기 일기예보 현황 처리 fucntion
    def shortcast_handler(self):

        # 배열 1st ==> 응답코드 + 응답메세지
        # 배열 2nd ==> items dict가 모여있는 list(없는 경우 비어 있는 list 값)
        # 배열 3rd ==> items 데이터 개수
        self.hold_flag = True
        self.conncet_db('초단기 일기예보 정보')
        list_col = ['UPLOAD_TIME', 'T1H', 'RN1', 'SKY', 'UUU', 'VVV', 'REH', 'PTY', 'LGT', 'VEC', 'WSD',
                    'NX', 'NY', 'STATE1', 'STATE2', 'STATE3', 'LOCATION', 'BASE_DATETIME', 'FCST_DATETIME']
        list_measure = ['T1H', 'RN1', 'SKY', 'UUU', 'VVV', 'REH', 'PTY', 'LGT', 'VEC', 'WSD']
        try:
            header = {'Content-Type': 'application/json; charset=utf-8'}
            current_time = self.get_currentTime()
            base_date_request = current_time[3][0]
            base_time_request = current_time[3][1]

            for idx, code in enumerate(self.local_code):
                # 각 변수 초기화
                response_data = []
                state1 = self.local_names[idx][0]
                state2 = self.local_names[idx][1]
                state3 = self.local_names[idx][2]
                location = self.city_names[idx]

                list_items = []
                list_fcst_tuples = []

                sql_insert = "REPLACE INTO shortcast("+', '.join(list_col)+") \
                VALUES(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                ########################################__실행해야 하는 지역코드 만큼 get api 실행 영역__########################################
                dict_parameter = 'serviceKey='+self.service_key+'&base_date='+base_date_request+'&base_time='+base_time_request+\
                                 '&nx='+code[0]+'&ny='+code[1]+'&numOfRows=300&pageNo=1&dataType=JSON'

                response = requests.get(self.shortcast_address, headers=header, params=dict_parameter)
                self.set_print('{} {} {} 단기일기예보 API 요청\n[REQUEST]\n{}'.format(state1, state2, state3,
                                                                              response.request.url))
                self.set_print('{} {} {} 단기일기예보 API 응답\n[RESPONSE]\nHeader:{}\nbody:{}'.format(state1, state2, state3,
                                                                               response.headers, response.text))
                # status_code 정상인 경우
                if response.status_code == 200:

                    response_json = response.json()
                    res_code = response_json['response']['header']['resultCode']
                    res_message = response_json['response']['header']['resultMsg']
                    # 정상 처리 됨
                    if res_code == '00':
                        items = response_json['response']['body']['items']['item']
                        data_size = response_json['response']['body']['totalCount']
                        response_data.append(res_code+' '+res_message)
                        response_data.append(items)
                        response_data.append(data_size)
                    # response code 에러 발생 비정상 처리
                    else:
                        raise Exception(
                            'REQUEST {} 에서 Error 발생 \n response code: {}\n response message: {}'.format
                            (response.request.url, res_code, res_message))
                # status_code 비정상의 경우
                else:
                    response.raise_for_status()
                ############################################__items parsing 작업 실행 영역__############################################
                return_base = self.change_format_bases(response_data[1][0]['baseDate'], response_data[1][0]['baseTime'])

                # list unique fcst_date and fcst_time data
                for item in response_data[1]:

                    fcst_data = (item['fcstDate'], item['fcstTime'])
                    if fcst_data not in list_fcst_tuples:
                        list_fcst_tuples.append(fcst_data)

                # mysql dict에 response data 넣어 주기
                for item in list_fcst_tuples:
                    creterion_fcst_date = item[0]
                    creterion_fcst_time = item[1]
                    return_fcst = self.change_format_bases(creterion_fcst_date, creterion_fcst_time)
                    mysql_items = {'UPLOAD_TIME': current_time[1], 'NX': code[0], 'NY': code[1], 'STATE1': state1,
                                   'STATE2': state2, 'STATE3': state3, 'BASE_DATETIME': return_base,
                                   'FCST_DATETIME': return_fcst, 'LOCATION': location}

                    for item2 in response_data[1]:

                        if item2['fcstDate'] == creterion_fcst_date and item2['fcstTime'] == creterion_fcst_time:
                            category = item2['category']
                            value = item2['fcstValue']
                            mysql_items[category] = str(value)

                    # data format change if need and check no data check
                    mysql_items['SKY'] = self.change_format_sky(mysql_items['SKY'])
                    mysql_items['PTY'] = self.change_format_pty(mysql_items['PTY'])
                    mysql_items['RN1'] = self.change_format_rn1(mysql_items['RN1'])
                    mysql_items['VEC'] = self.change_format_vec(mysql_items['VEC'])
                    mysql_items['VVV'] = self.change_format_vvv(mysql_items['VVV'])
                    mysql_items['UUU'] = self.change_format_uuu(mysql_items['UUU'])
                    mysql_items['LGT'] = self.change_format_lgt(mysql_items['LGT'])


                    ############################################__MySQL DB에 Upload 데이터 실행 영역__###########################################
                    # mysql tuple mysql insert 순서로 데이터 생성
                    list_temp_mysql = []
                    for col in list_col:
                        list_temp_mysql.append(mysql_items[col])
                    mysql_tuple = tuple(list_temp_mysql)
                    list_items.append(mysql_tuple)

                # mysql 업로드 진행
                self.set_print('{} {} {} DB 단기일기예보 정보 {}건 업로드 진행'.format(state1, state2, state3, len(list_items)))
                for tuple_data in list_items:
                    self.cur.execute(sql_insert, tuple_data)
                self.conn.commit()
                self.set_print('{} {} {} DB 단기일기예보 정보 업로드 완료'.format(state1, state2, state3))

        except:
            self.set_print('Error: {}. {}, line: {}'.format(sys.exc_info()[0], sys.exc_info()[1], sys.exc_info()[2].tb_lineno))

        finally:
            self.set_print('초단기 일기예보 정보 DB 연결 해제')
            self.cur.close()
            self.conn.close()
            self.hold_flag = False

    # 동네 현황 처리 function
    def weather_handler(self):

        # 배열 1st ==> 응답코드+응답메세지
        # 배열 2nd ==> items dict가 모여있는 list(없는 경우 비어 있는 list 값)
        # 배열 3rd ==> items 데이터 개수
        self.hold_flag = True
        self.conncet_db('지역 기상 정보')
        try:
            header = {'Content-Type': 'application/json; charset=utf-8'}
            current_time = self.get_currentTime()
            base_date_request = current_time[3][0]
            base_time_request = current_time[3][1]

            for idx, code in enumerate(self.local_code):
                # 각 변수 초기화
                response_data = []
                state1 = self.local_names[idx][0]
                state2 = self.local_names[idx][1]
                state3 = self.local_names[idx][2]
                location = self.city_names[idx]

                mysql_items = {'UPLOAD_TIME': current_time[1], 'NX': code[0], 'NY': code[1], 'STATE1': state1,
                               'STATE2': state2, 'STATE3': state3, 'LOCATION': location}
                list_col = ['UPLOAD_TIME', 'T1H', 'RN1', 'UUU', 'VVV', 'REH', 'PTY', 'VEC', 'WSD', 'NX', 'NY',
                            'STATE1', 'STATE2', 'STATE3', 'LOCATION', 'BASE_DATETIME']

                sql_insert = "REPLACE INTO weather("+', '.join(list_col)+") \
                VALUES(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                ########################################__실행해야 하는 지역코드 만큼 get api 실행 영역__########################################
                dict_parameter = 'serviceKey='+self.service_key+'&base_date='+base_date_request+'&base_time='+base_time_request+\
                                 '&nx='+code[0]+'&ny='+code[1]+'&numOfRows=300&pageNo=1&dataType=JSON'

                response = requests.get(self.weather_address, headers=header, params=dict_parameter)
                self.set_print('{} {} {} 단기날씨실황 API 요청\n[REQUEST]\n{}'.format(state1, state2, state3,
                                                                              response.request.url))
                self.set_print('{} {} {} 단기날씨실황 API 응답\n[RESPONSE]\nHeader:{}\nbody:{}'.format(state1, state2, state3,
                                                                               response.headers, response.text))
                # status_code 정상인 경우
                if response.status_code == 200:

                    response_json = response.json()
                    res_code = response_json['response']['header']['resultCode']
                    res_message = response_json['response']['header']['resultMsg']
                    # 정상 처리 됨
                    if res_code == '00':
                        items = response_json['response']['body']['items']['item']
                        data_size = response_json['response']['body']['totalCount']
                        response_data.append(res_code+' '+res_message)
                        response_data.append(items)
                        response_data.append(data_size)
                    # response code 에러 발생 비정상 처리
                    else:
                        raise Exception(
                            'REQUEST {} 에서 Error 발생 \n response code: {}\n response message: {}'.format
                            (response.request.url, res_code, res_message))
                # status_code 비정상의 경우
                else:
                    response.raise_for_status()
                ############################################__items parsing 작업 실행 영역__############################################
                # mysql dict에 response data 넣어 주기
                for idx2, item in enumerate(response_data[1]):
                    category = item['category']
                    value = item['obsrValue']
                    if idx2 == 0:
                        return_base = self.change_format_bases(item['baseDate'], item['baseTime'])
                        mysql_items['BASE_DATETIME'] = return_base

                    mysql_items[category] = str(value)

                # data format change if need
                mysql_items['RN1'] = self.change_format_rn1(mysql_items['RN1'])
                mysql_items['PTY'] = self.change_format_pty(mysql_items['PTY'])
                mysql_items['VEC'] = self.change_format_vec(mysql_items['VEC'])
                mysql_items['VVV'] = self.change_format_vvv(mysql_items['VVV'])
                mysql_items['UUU'] = self.change_format_uuu(mysql_items['UUU'])

                ############################################__MySQL DB에 Upload 데이터 실행 영역__###########################################
                # mysql tuple mysql insert 순서로 데이터 생성
                list_temp_mysql = []
                for col in list_col:
                    list_temp_mysql.append(mysql_items[col])
                mysql_tuple = tuple(list_temp_mysql)
                self.set_print('{} {} {} DB 날씨실항 정보 업로드 진행'.format(state1, state2, state3))
                self.cur.execute(sql_insert, mysql_tuple)
                self.conn.commit()
                self.set_print('{} {} {} DB 날씨실항 정보 업로드 완료'.format(state1, state2, state3))

        except:
            self.set_print('Error Occurred : {}. {}, line: {}'.format(sys.exc_info()[0], sys.exc_info()[1], sys.exc_info()[2].tb_lineno))

        finally:
            self.set_print('지역 기상 정보 DB 연결 해제')
            self.cur.close()
            self.conn.close()
            self.hold_flag = False

    # 일몰/일출 처리 function
    def riseset_handler(self):

        # 배열 1st ==> 응답코드+응답메세지
        # 배열 2nd ==> items dict가 모여있는 list(없는 경우 비어 있는 list 값)
        # 배열 3rd ==> items 데이터 개수
        self.hold_flag = True
        self.conncet_db('일출몰/월출몰 정보')
        try:
            header = {'Content-Type': 'text/xml; charset=utf-8'}
            list_col = ['UPLOAD_TIME', 'LOCDATE', 'LOCATION', 'LONGITUDE', 'LONGITUDE_NUM', 'LATITUDE',
                        'LATITUDE_NUM', 'SUNRISE', 'SUNTRANSIT', 'SUNSET', 'MOONRISE', 'MOONTRANSIT', 'MOONSET',
                        'CIVILM', 'CIVILE', 'NAUTM', 'NAUTE', 'ASTM', 'ASTE']

            list_key = ['locdate', 'location', 'longitude', 'longitudeNum', 'latitude', 'latitudeNum', 'sunrise',
                        'suntransit', 'sunset', 'moonrise', 'moontransit', 'moonset', 'civilm', 'civile', 'nautm',
                        'naute', 'astm', 'aste']

            sql_insert = "REPLACE INTO riseset(" + ', '.join(list_col) + ") \
            VALUES(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"

            current_time = self.get_currentTime()
            base_date_request = current_time[3][0]

            # DB local에 있는 city name 얻어오기
            for city in self.city_names:

                # 각 변수 초기화
                recent_date = self.get_riseset_recent_date(city)
                list_base_date = None
                mysql_items = {'UPLOAD_TIME': current_time[1]}

                # riseset의 DB에 정보가 있어서 최신 Date가 존재 하는 경우 해당 최신일을 포함하지 않는 +5일 데이터를 가져온다
                if recent_date:
                    recent_date_str = recent_date.strftime("%Y%m%d")
                    recent_date_reform = datetime.strptime(recent_date_str, "%Y%m%d").date()
                    today_date_reform = datetime.strptime(base_date_request, "%Y%m%d").date()
                    between_date = recent_date_reform - today_date_reform
                    diff_days = between_date.days

                    # 만약 서버에 당일보다 전에 날짜로 recent date일 경우(정상적이지 않을 경우)
                    # server에 date 정보가 없는 것과 동일하게 동작(recent_date=None 과 동일 시)
                    if diff_days <= 0:
                        list_base_date = self.get_riseset_list_date(base_date_request, include=True)
                    # 이미 recent_date가 최신인 경우 skip 처리
                    elif diff_days >= 4:
                        self.set_print('{} 지역 riseset 정보 이미 {} 오늘 기준 +4일의 최신 정보로 되어 있어 SKIP 처리'
                                       .format(city, recent_date))
                        continue
                    # recent_date이 금일 기준으로(불포함) +5일이 되도록 유지함
                    else:
                        list_base_date = self.get_riseset_list_date(recent_date_str, include=False, count=(4-diff_days))

                # riseset의 DB에 정보가 없는 경우 해당 최신일을 포함하는 않는 +5일 데이터를 가져온다
                else:
                    list_base_date = self.get_riseset_list_date(base_date_request, include=True)

                self.set_print('{} 지역 riseset 기준날짜 List: {}'.format(city, list_base_date))

                # list_base_date 날짜만큼 일출몰/월출몰 데이터 획득(기본 5일)
                for base_date in list_base_date:
                    #  connect DB 및 각 변수 초기화
                    response_data = []
                    ########################################__실행해야 하는 날짜 만큼 riseset get api 실행 영역__########################################
                    dict_parameter = 'serviceKey='+self.service_key_riseset+'&locdate='+base_date+'&location='+city
                    # 해당 API는 reponse가 xml로 내려옴(xml to json dict로 변환이 필요)
                    response = requests.get(self.riseset_address, headers=header, params=dict_parameter)
                    self.set_print('{} 일출몰/월출몰 API 요청\n[REQUEST]\n{}'.format(city, response.request.url))
                    self.set_print('{} 일출몰/월출몰 API 응답\n[RESPONSE]\nHeader:{}\nbody:{}'.format(city, response.headers, response.text))

                    # status_code 정상인 경우
                    if response.status_code == 200:

                        xml_string = response.text
                        # xml to json type
                        response_json = json.dumps(xmltodict.parse(xml_string))
                        # json to dict type for python
                        response_json = json.loads(response_json)
                        # result code and result msg
                        res_code = response_json['response']['header']['resultCode']
                        res_message = response_json['response']['header']['resultMsg']
                        # 정상 처리 됨
                        if res_code == '00':

                            items = response_json['response']['body']['items']['item']
                            data_size = response_json['response']['body']['totalCount']
                            response_data.append(res_code+' '+res_message)
                            response_data.append(items)
                            response_data.append(data_size)
                        # response code 에러 발생 비정상 처리
                        else:
                            raise Exception(
                                'REQUEST {} 에서 Error 발생 \n response code: {}\n response message: {}'.format
                                (response.request.url, res_code, res_message))
                    # status_code 비정상의 경우
                    else:
                        response.raise_for_status()

                    ############################################__items parsing 작업 실행 영역__############################################
                    # mysql dict에 response data 넣어 주기
                    for idx, key in enumerate(list_key):

                        if '-' in response_data[1][key]:
                            mysql_items[list_col[idx+1]] = '없음'
                            continue
                        if key == 'locdate':
                            changes = self.change_format_bases(response_data[1][key], '0000')
                            changes = changes.split()
                            mysql_items[list_col[idx+1]] = changes[0].strip()
                            continue

                        mysql_items[list_col[idx+1]] = response_data[1][key]

                    ############################################__MySQL DB에 Upload 데이터 실행 영역__###########################################
                    # mysql tuple mysql insert 순서로 데이터 생성
                    list_temp_mysql = []
                    for col in list_col:
                        list_temp_mysql.append(mysql_items[col])
                    mysql_tuple = tuple(list_temp_mysql)

                    self.set_print('{} {} DB 일출몰/월출몰 정보 업로드 진행'.format(base_date, city))
                    self.cur.execute(sql_insert, mysql_tuple)
                    self.conn.commit()
                    self.set_print('{} {} DB 일출몰/월출몰 정보 업로드 완료'.format(base_date, city))

        except:
            self.set_print('Error Occurred : {}. {}, line: {}'.format(sys.exc_info()[0], sys.exc_info()[1], sys.exc_info()[2].tb_lineno))

        finally:
            self.set_print('일출몰/월출몰 정보 DB 연결 해제')
            self.cur.close()
            self.conn.close()
            self.hold_flag = False

    # BackgroundScheduler에 타입(cron / interval)에 따른 스케쥴 등록
    def setting_jod(self, forecast_data, weather_data, shortcast_data, riseset_data):

        # weather_schedule job setting
        # scheduler type이 cron 인 경우
        if weather_data[0] == 'cron':
            # weather_schedule 에서 시간 데이터가 없을 경우 매 시간 42분 실행
            if weather_data[1] == '':
                self.sched_module.add_job(self.weather_handler, 'cron', args=None, hour="0-23", minute="42",
                                          id='weather_cron_one_hour', misfire_grace_time=360, replace_existing=False,
                                          coalesce=True)
            # weather_schedule 에서 시간 데이터가 있을 경우 해당 시간에 실행
            else:
                for idx, time in enumerate(weather_data[1]):
                    temp_time = str(time).split(":")
                    temp_hour = temp_time[0]
                    temp_minute = temp_time[1]
                    # 시간 데이터 (분) 가공
                    if temp_minute == '00':
                        temp_minute = '0'
                    elif temp_minute[0] == '0':
                        temp_minute = temp_minute[1:]
                    self.sched_module.add_job(self.weather_handler, 'cron', args=None, hour=temp_hour, minute=temp_minute,
                                              id='weather_cron_' + str(idx), misfire_grace_time=360,
                                              replace_existing=False, coalesce=True)
        # scheduler type이 interval인 경우
        else:
            # interval 값이 비어 있는 경우 매 시간마다 실행
            if weather_data[2] == '':
                self.sched_module.add_job(self.weather_handler, 'interval', args=None, hours=1, id='weather_interval_one_hour',
                                          misfire_grace_time=360, replace_existing=False, coalesce=True)
            # weather_schedule 에서 interval 있을 경우 해당 시간에 실행
            else:
                interval_value = int(weather_data[2])
                self.sched_module.add_job(self.weather_handler, 'interval', args=None, minutes=interval_value,
                                          id='weather_interval_'+weather_data[2],
                                          misfire_grace_time=360, replace_existing=False, coalesce=True)

        # shortcast_schedule job setting
        # scheduler type이 cron 인 경우
        if shortcast_data[0] == 'cron':
            # weather_schedule 에서 시간 데이터가 없을 경우 매 시간 42분 실행
            if shortcast_data[1] == '':
                self.sched_module.add_job(self.shortcast_handler, 'cron', args=None, hour="0-23", minute="42",
                                          id='shortcast_cron_one_hour', misfire_grace_time=360, replace_existing=False,
                                          coalesce=True)
            # weather_schedule 에서 시간 데이터가 있을 경우 해당 시간에 실행
            else:
                for idx, time in enumerate(shortcast_data[1]):
                    temp_time = str(time).split(":")
                    temp_hour = temp_time[0]
                    temp_minute = temp_time[1]
                    # 시간 데이터 (분) 가공
                    if temp_minute == '00':
                        temp_minute = '0'
                    elif temp_minute[0] == '0':
                        temp_minute = temp_minute[1:]
                    self.sched_module.add_job(self.shortcast_handler, 'cron', args=None, hour=temp_hour, minute=temp_minute,
                                              id='shortcast_cron_' + str(idx), misfire_grace_time=360,
                                              replace_existing=False, coalesce=True)
        # scheduler type이 interval인 경우
        else:
            # interval 값이 비어 있는 경우 매 시간마다 실행
            if shortcast_data[2] == '':
                self.sched_module.add_job(self.shortcast_handler, 'interval', args=None, hours=1, id='shortcast_interval_one_hour',
                                          misfire_grace_time=360, replace_existing=False, coalesce=True)
            # weather_schedule 에서 interval 있을 경우 해당 시간에 실행
            else:
                interval_value = int(shortcast_data[2])
                self.sched_module.add_job(self.shortcast_handler, 'interval', args=None, minutes=interval_value,
                                          id='shortcast_interval_'+shortcast_data[2],
                                          misfire_grace_time=360, replace_existing=False, coalesce=True)

        # forecast_schedule job setting
        # scheduler type이 cron 인 경우
        if forecast_data[0] == 'cron':
            # forecast_schedule 에서 시간 데이터가 없을 경우 매 시간 42분 실행
            if forecast_data[1] == '':
                self.sched_module.add_job(self.forecast_handler, 'cron', args=None, hour="0-23", minute="12",
                                          id='forecast_cron_one_hour', misfire_grace_time=360, replace_existing=False,
                                          coalesce=True)
            # weather_schedule 에서 시간 데이터가 있을 경우 해당 시간에 실행
            else:
                for idx, time in enumerate(forecast_data[1]):
                    temp_time = str(time).split(":")
                    temp_hour = temp_time[0]
                    temp_minute = temp_time[1]
                    # 시간 데이터 (분) 가공
                    if temp_minute == '00':
                        temp_minute = '0'
                    elif temp_minute[0] == '0':
                        temp_minute = temp_minute[1:]
                    self.sched_module.add_job(self.forecast_handler, 'cron', args=None, hour=temp_hour, minute=temp_minute,
                                              id='forecast_cron_' + str(idx), misfire_grace_time=360,
                                              replace_existing=False, coalesce=True)
        # scheduler type이 interval인 경우
        else:
            # interval 값이 비어 있는 경우 매 시간마다 실행
            if forecast_data[2] == '':
                self.sched_module.add_job(self.forecast_handler, 'interval', args=None, hours=1, id='forecast_interval_one_hour',
                                          misfire_grace_time=360, replace_existing=False, coalesce=True)
            # weather_schedule 에서 interval 있을 경우 해당 시간에 실행
            else:
                interval_value = int(forecast_data[2])
                self.sched_module.add_job(self.forecast_handler, 'interval', args=None, minutes=interval_value,
                                          id='forecast_interval_'+forecast_data[2],
                                          misfire_grace_time=360, replace_existing=False, coalesce=True)

        # riseset_schedule job setting
        # scheduler type이 cron 인 경우
        if riseset_data[0] == 'cron':
            # weather_schedule 에서 시간 데이터가 없을 경우 매 시간 42분 실행
            if riseset_data[1] == '':
                self.sched_module.add_job(self.riseset_handler, 'cron', args=None, hour="0-23", minute="42",
                                          id='riseset_cron_one_hour', misfire_grace_time=360, replace_existing=False,
                                          coalesce=True)
            # weather_schedule 에서 시간 데이터가 있을 경우 해당 시간에 실행
            else:
                for idx, time in enumerate(riseset_data[1]):
                    temp_time = str(time).split(":")
                    temp_hour = temp_time[0]
                    temp_minute = temp_time[1]
                    # 시간 데이터 (분) 가공
                    if temp_minute == '00':
                        temp_minute = '0'
                    elif temp_minute[0] == '0':
                        temp_minute = temp_minute[1:]
                    self.sched_module.add_job(self.riseset_handler, 'cron', args=None, hour=temp_hour, minute=temp_minute,
                                              id='riseset_cron_' + str(idx), misfire_grace_time=360,
                                              replace_existing=False, coalesce=True)
        # scheduler type이 interval인 경우
        else:
            # interval 값이 비어 있는 경우 매 시간마다 실행
            if riseset_data[2] == '':
                self.sched_module.add_job(self.riseset_handler, 'interval', args=None, hours=1, id='riseset_interval_one_hour',
                                          misfire_grace_time=360, replace_existing=False, coalesce=True)
            # weather_schedule 에서 interval 있을 경우 해당 시간에 실행
            else:
                interval_value = int(riseset_data[2])
                self.sched_module.add_job(self.riseset_handler, 'interval', args=None, minutes=interval_value,
                                          id='riseset_interval_'+riseset_data[2],
                                          misfire_grace_time=360, replace_existing=False, coalesce=True)

    # main executes function
    def run(self):
        try:
            self.introText = """
            #################################################-WEATHER UPLOADER-#############################################

                1. Weather API 연동 스케쥴러 프로그램 입니다.
                2. 공공데이터 날씨 API인 동네예보 조회서비스 제공입니다. 
                3. 시간 및 간격 Config 정보 파일을(GET_WEATHER_CONFIG.xlsx) 프로그램이 위치한 경로에 준비하셔야 합니다.
                4. Config 파일에 '지역코드표, 지역설정, 일기예보시간설정, 단기실황시간설정, 주소정보, DB정보' Sheet가 있어야 합니다.
                7. Config 파일의 'Time'에 '일기예보시간설정, 단기실황시간설정'을 명시하면 특정 시간에 동작하게 됩니다.(Optional)
                8. Interval 설정 시 분단위로 입력입니다.
                9. Config 파일의 'DB정보'에 아이디/비밀번호 등의 정보를 명시해야 합니다.

            ##############################################################################################################\n\n
            """
            print(self.introText)
            # Check File exists
            self.config_flag = os.path.isfile(self.config_path)

            # 프로그램 진행 Process logic
            if self.config_flag:
                self.flag = input("Config 파일이 준비된 상태로 지금 프로그램을 진행 하시겠습니까?(y|n) : ")
                self.flag.lower()
                # check intro flag
                if self.flag == 'y':
                    self.set_print("설정 파일 경로: {}".format(self.config_path))
                    pass
                elif self.flag == 'n':
                    self.set_print("프로그램을 종료합니다.")
                    sys.exit(0)
                else:
                    while self.flag is not "y" and self.flag is not "n":
                        self.flag = input("잘못입력하셨습니다. \"y\" 또는 \"n\"을 입력하여 주세요. : ")
                        self.flag.lower()
                        if self.flag == 'n':
                            self.set_print("프로그램을 종료합니다.")
                            sys.exit(0)
                        elif self.flag == 'y':
                            break
                        else:
                            pass
            else:
                self.set_print("현재 프로그램이 위치한 경로에 'GET_WEATHER_CONFIG.xlsx' 설정 엑셀 파일을 위치해 주시고 다시 시작해주세요.")
                sys.exit(0)

            # scheduler 생성 및 job 생성
            self.sched_module = BackgroundScheduler()
            self.sched_module.start()
            # Config Data setting
            self.get_configFile()
            # job generate
            # self.setting_jod(self.type, [self.cron_time, self.interval])
            self.setting_jod(self.forecast_schedules, self.weather_schedules, self.shortcast_schedules,
                             self.riseset_schedules)

            # start
            while True:
                if not self.hold_flag:
                    self.set_print("Get Weather Program is Running now...")
                sleep(60)
        except:
            self.set_print('Error Occurred : {}. {}, line: {}'.format(sys.exc_info()[0], sys.exc_info()[1], sys.exc_info()[2].tb_lineno))
            sys.exit(-1)


if __name__ == "__main__":
    gw = GetWeather()
    gw.run()
