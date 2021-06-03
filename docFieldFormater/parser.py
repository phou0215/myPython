# -*- coding: utf-8 -*-
"""
Created on Thr Oct 18 14:25:12 2019
@author: TestEnC hanrim lee

"""
import os
import sys
import re
import threading
import openpyxl
import easygui

from os.path import expanduser
from time import sleep
from datetime import datetime

from openpyxl.styles import Alignment, Font, NamedStyle, PatternFill
from openpyxl import formatting, styles, Workbook
from openpyxl.styles.borders import Border, Side
# EARFCN 275 - BAND1 , EARFCN 1350 - BAND3 , EARFCN 2500 - BAND5, EARFCN 2850 - BAND7, EARFCN 3200 - BAND8


class DocFormatter():
    # ####################################################__초기화__####################################################

    def __init__(self):

        super(DocFormatter, self).__init__()
        self.input_path = None
        self.output_path = ""
        self.dict_cell_info = {}
        self.workbook_input = None
        self.workbook_output = None
        self.worksheet_ps = None
        self.worksheet_sum = None
        self.layer_1 = None
        self.layer_2 = None
        self.layer_3 = None
        self.layer_4 = None
        self.layer_5 = None
        self.total_layer = None
        self.total_rows = 0
        self.list_combination = None
        self.list_col_layer = ['A', 'B', 'C', 'D', 'E']
        self.list_col_summary = ['F', 'G', 'H', 'I', 'J']
        self.dict_band = {'275': 'BAND1', '1350': 'BAND3', '2500': 'BAND5', '2850': 'BAND7', '3200': 'BAND8'}

        # Pattern_Fill
        self.gray_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        self.brown_fill = PatternFill(start_color='DDD9C4', end_color='DDD9C4', fill_type='solid')
        self.dark_gray_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        # style font color and size
        self.red_font = Font(name='맑은 고딕', size=10, bold=True, color='FF0000')
        self.blue_font = Font(name='맑은 고딕', size=10, bold=True, color='0000FF')
        self.index_font = Font(name='맑은 고딕', size=11, bold=True, color='2B2B2B')
        # style Alignment
        self.general_alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        self.top_alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")
        # # style border
        self.thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # #################################################################################################################

    # ###################################################__Util 함수__##################################################

    # 로그 문자 처리 함수
    def setPrint(self, text):

        str_today = datetime.today().strftime("%Y-%m-%d %H:%M:%S")
        print(str_today+":\n" + text + "\n")

    # 현재 시간 구하여 3가지 타입으로 list return
    def getCurrent_time(self):

        nowTime = datetime.now()
        now_datetime_str = nowTime.strftime('%Y-%m-%d %H:%M:%S')
        now_datetime_str2 = nowTime.strftime('%Y-%m-%d %H%M%S')
        return [nowTime, now_datetime_str, now_datetime_str2]

    # 문자열 특수문자 제거 함수
    def remove_special(self, text):

        return_text = re.sub('[-=+,#/\?^$@*\"※~&%ㆍ!』\\‘|\(\)\[\]\<\>\{\}`><]\'', '', text)
        return return_text

    # 문장 앞 부터 조건에 맞는 문자열 substring
    def find_between(self, s, first, last):
        try:
            start = s.index(first)+len(first)
            end = s.index(last, start)
            return_data = s[start:end]
            return return_data
        except ValueError:
            return return_data

    # 문장 뒤 부터 조건에 맞는 문자열 substring
    def find_between_r(self, s, first, last):
        try:
            return_data = ""
            start = s.rindex(first)+len(first)
            end = s.rindex(last, start)
            return_data = s[start:end]
            return return_data
        except ValueError:
            return return_data

    # check convert num available
    def check_number(self, string):

        try:
            float(string)
            return True
        except:
            return False

    # extract condition in extract sheet 'A', 'B', 'C'... columns
    def extract_data(self, ws, list_col):

        list_return = []
        for item in list_col:
            list_temp = []
            for cell in ws[item]:
                cell_value = str(cell.value).replace(" ", "")
                list_temp.append(cell_value)
            list_return.append(list_temp)
        return list_return

    # check convert num available
    def check_empty(self, string_data):

        return_data = None
        if string_data is None or string_data == '' or string_data.lower() in ['n/a', 'na', 'nt', 'n/t']:
            return_data = '-'
        else:
            return_data = str(string_data)

        return return_data

    def check_band(self, idx, list_band_info):

        try:
            return_check = ['OK', '']
            if self.layer_1[idx] != '2850':
                return_check = ['대상아님', '']
                return return_check
            else:
                for idx2, band in enumerate(list_band_info):

                    # check double band in layer info
                    layer_freq = self.total_layer[idx2][idx]
                    if ";" in layer_freq:
                        list_freq_temp = layer_freq.split(";")
                        measure_band = ""
                        for freq in list_freq_temp:
                            if band.strip().lower() == self.dict_band[freq].lower():
                                measure_band = self.dict_band[freq]
                                break

                        if measure_band == "":
                            return_check[0] = 'FAULT'
                            return_check[1] = return_check[1] + '-' + layer_freq + ' 에 ' + band + '가 포함되지 않음\r\n'
                        else:
                            continue
                    # single band info
                    else:
                        measure_band = self.dict_band[layer_freq]
                        if band.strip().lower() != measure_band.strip().lower():
                            return_check[0] = 'FAULT'
                            return_check[1] = return_check[1] + '-' + band + ' 와 ' + measure_band + '불일치\r\n'
                        else:
                            continue
            return return_check

        except:
            self.setPrint('Error: {}. {}, line: {}'.format(
                sys.exc_info()[0],
                sys.exc_info()[1],
                sys.exc_info()[2].tb_lineno) + ' ')

    def generate_output_dir(self):

        try:
            root_path = expanduser('~')
            sub_path = root_path + "\\Desktop\\Field\\"
            dir_exist = os.path.isdir(sub_path)
            if not dir_exist:
                os.mkdir(sub_path)
            current_time = self.getCurrent_time()[2]
            os.mkdir(sub_path + 'result_' + current_time + '\\')
            self.output_path = sub_path + 'result_' + current_time + '\\' + 'result_' + current_time + ".xlsx"

        except:
            self.setPrint('Error: {}. {}, line: {}'.format(
                sys.exc_info()[0],
                sys.exc_info()[1],
                sys.exc_info()[2].tb_lineno) + ' ')

    # #################################################################################################################

    # generate ca file data sheet and save
    def generate_sheet(self):

        try:
            self.setPrint('분석 대상 파일 경로: {}'.format(self.input_path))
            self.workbook_input = openpyxl.load_workbook(self.input_path, data_only=True)
            list_sheet_names = self.workbook_input.sheetnames
            list_sheet_names = [name.strip() for name in list_sheet_names if name]
            # 모든 필요 시트 충족 조건
            if 'PS Call Statistics' not in list_sheet_names:
                self.setPrint('분석 가능한 \"PS Call Statistics Sheet\" 가 없습니다. 입력 파일을 확인해 주세요')
                sys.exit(-1)

            self.workbook_output = self.workbook_input
            self.worksheet_ps = self.workbook_output['PS Call Statistics']
            # self.worksheet_sum = self.workbook_output.copy_worksheet(self.worksheet_ps)
            self.worksheet_sum = self.workbook_output.create_sheet('Summary')

            # max_row max_column
            self.total_rows = self.worksheet_ps.max_row

            # A, B, C column 조건 값 추출
            self.total_layer = self.extract_data(self.worksheet_ps, ['H', 'I', 'J', 'K', 'L', 'O'])

            self.layer_1 = self.total_layer[0]
            self.layer_2 = self.total_layer[1]
            self.layer_3 = self.total_layer[2]
            self.layer_4 = self.total_layer[3]
            self.layer_5 = self.total_layer[4]
            self.list_combination = self.total_layer[5]

        except:
            self.setPrint('Error: {}. {}, line: {}'.format(
                sys.exc_info()[0],
                sys.exc_info()[1],
                sys.exc_info()[2].tb_lineno)+' ')

    # summary Tab
    def summary_generate_data(self):

        try:
            for idx, item in enumerate(self.layer_1):

                # #######################__layer column data input__#######################
                for idx2, col in enumerate(self.list_col_layer):
                    # get data from wb_input
                    self.worksheet_sum[col+str(idx+1)] = self.total_layer[idx2][idx]
                # #########################################################################

                # #######################__summary column data input__#####################
                # input column name
                if idx == 0:
                    self.worksheet_sum.merge_cells('F1:J1')
                    self.worksheet_sum['F1'] = self.list_combination[0]
                    self.worksheet_sum['K1'] = 'BAND_ACCURACY'
                    self.worksheet_sum['L1'] = 'ERROR_INFO'
                else:
                    band_info = self.list_combination[idx]
                    list_band_info = band_info.split("+")
                    band_size = len(list_band_info)
                    matching = self.check_band(idx, list_band_info)
                    for idx2, col in enumerate(self.list_col_summary[:band_size]):
                        self.worksheet_sum[col+str(idx+1)] = list_band_info[idx2]
                    self.worksheet_sum['K'+str(idx+1)] = matching[0]
                    self.worksheet_sum['L'+str(idx+1)] = matching[1]
                # #########################################################################

            # all cell alignment adjust
            for idx, mCell in enumerate(self.worksheet_sum["A1:L" + str(self.total_rows)]):
                for cell in mCell:
                    cell.alignment = self.general_alignment
                    if cell.column == 11:
                        cell_val = cell.value
                        if cell_val == 'OK':
                            cell.font = self.blue_font
                        elif cell_val == 'FAULT':
                            cell.font = self.red_font
                        else:
                            pass

            # each column width adjust
            sheet_cell_list = ['A', 'B', 'C', 'D', 'E',
                               'F', 'G', 'H', 'I', 'J',
                               'K', 'L']
            sheet_width_list = [32.75, 48.88, 48.88, 48.88, 48.88,
                                10.25, 10.25, 10.25, 10.25, 10.25,
                                20.25, 37.88]
            for i in range(len(sheet_cell_list)):
                self.worksheet_sum.column_dimensions[sheet_cell_list[i]].width = sheet_width_list[i]

            # Each Sheet set auto filter column
            self.worksheet_sum.auto_filter.ref = "A1:L" + str(self.total_rows)

            # Set cell styles
            col_list = self.list_col_layer + self.list_col_summary
            col_list = col_list + ['K', 'L']
            for col in col_list:
                self.worksheet_sum[col + '1'].font = self.index_font
                self.worksheet_sum[col + '1'].fill = self.brown_fill
                self.worksheet_sum[col + '1'].border = self.thin_border

            # save file
            self.workbook_output.save(self.output_path)
        except:
            self.setPrint('Error: {}. {}, line: {}'.format(sys.exc_info()[0], sys.exc_info()[1], sys.exc_info()[2].tb_lineno)+' ')

    # main method
    def run(self):

        try:
            self.introText = """
            ############################################-SMAPLE FILELD DOC FORMAT-#########################################
                1. FIELD 팀 BAND 정보 분석 Demo 도구 입니다.
                2. 여기에 내용을 쓰세요.
            ##############################################################################################################\n\n
            """
            print(self.introText)

            # 프로그램 진행 Process logic
            flag = input("Config 파일이 준비된 상태로 지금 프로그램을 진행 하시겠습니까?(y|n) : ")
            flag.lower()
            # check intro flag
            if flag == 'y':
                pass
            elif flag == 'n':
                self.setPrint("프로그램을 종료합니다.")
                sys.exit(-1)
            else:
                while flag is not "y" and flag is not "n":
                    flag = input("잘못입력하셨습니다. \"y\" 또는 \"n\"을 입력하여 주세요. : ")
                    flag.lower()
                    if flag == 'n':
                        self.setPrint("프로그램을 종료합니다.")
                        sys.exit(-1)
                    elif flag == 'y':
                        break
                    else:
                        pass

            # select File
            filetypes = ["*.xlsx"]
            self.input_path = easygui.fileopenbox(default='*', filetypes=filetypes, multiple=False)
            while not self.input_path:
                self.setPrint('선택된 파일이 없습니다. 다시 선택해 주세요.')
                self.input_path = easygui.fileopenbox(default='*', filetypes=filetypes, multiple=False)

            start_time = self.getCurrent_time()[1]
            self.setPrint("STARTED_TIME: " + start_time + " ")

            # root and sub directory generate
            self.generate_output_dir()

            # get target band data
            self.generate_sheet()

            # core function start
            self.summary_generate_data()

            # Core Code
            end_time = self.getCurrent_time()[1]
            # Excel input Data read
            self.setPrint("FINISHED_TIME: "+end_time+" ")

        except:
            self.setPrint('Error: {}. {}, line: {}'.format(sys.exc_info()[0],
                                                              sys.exc_info()[1],
                                                              sys.exc_info()[2].tb_lineno)+' ')


if __name__ == '__main__':
    moduler = DocFormatter()
    moduler.run()
