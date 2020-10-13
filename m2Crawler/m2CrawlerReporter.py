#-*- coding:utf-8 -*-

import smtplib
import os
import sys
import pandas as pd
from datetime import datetime
# Excel openpyxl library
import openpyxl

from openpyxl.styles import Alignment, Font, NamedStyle, PatternFill
from openpyxl import formatting, styles, Workbook
from openpyxl.styles.borders import Border, Side


class ResultReport():

    # 클래스 초기화
    def __init__(self, report_path, list_dict_result):

        super().__init__()
        self.list_cols = ['bo_no', 'cafeName', 'nav', 'type', 'deviceName', 'reportDate', 'reportTime', 'regiDate', 'title',
                          'story', 'reply', 'cate', 'status', 'reply_num', 'view_num', 'story_url', 'img_url', 'video_url', 'uniqueId']
        self.list_cols_2 = ["게시글 번호", "카페이름", "네비게이션", "구분", "디바이스", "작성일", "작성시간", "등록일", "제목", "내용",
                            "댓글", "분류", "상태", "댓글수", "조회수", "게시글 주소", "이미지 주소", "동영상 주소", "고유번호"]
        self.report_path = report_path
        self.report_final_path = ''
        self.list_dict_result = list_dict_result
        # 채우기 스타일
        self.blue_fill = PatternFill(start_color='DDD9C4', end_color='DDD9C4', fill_type='solid')

        # style font color and size
        self.top_font = Font(name='맑은 고딕', size=11, bold=True, color='2B2B2B')
        self.value_normal_font = Font(name='맑은 고딕', size=10, bold=False, color='2B2B2B')
        self.value_bold_font = Font(name='맑은 고딕', size=10, bold=True, color='2B2B2B')

        # style Alignment
        self.top_alignment = Alignment(wrap_text=False, horizontal="center", vertical="center")
        self.general_alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")

        # style border
        self.thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # 시간 정보와 함께 Console에 Print 하는 함수
    def setPrint(self, text):

        self.current = datetime.today().strftime("%Y-%m-%d %H:%M:%S")
        print("{}:\n{}".format(self.current, text)+"\n")

    # 현재 시간 구하여 3가지 타입으로 list return
    def getCurrent_time(self):

        self.nowTime = datetime.now()
        self.nowtime_str = self.nowTime.strftime('%Y-%m-%d %H:%M:%S')
        self.nowtime_str_2 = self.nowTime.strftime('%Y-%m-%d %H %M %S')
        return [self.nowTime, self.nowtime_str, self.nowtime_str_2]

    # adjusting excel options
    def set_excelOp(self, list_keys, list_count_rows, save_path):

        wb = openpyxl.load_workbook(save_path, data_only=True)
        for i, key in enumerate(list_keys):

            tot_rows = list_count_rows[i]+1
            sheet = wb[key]
            # date_style = NamedStyle(name="datetime", number_format="YYYY-MM-DD")

            list_col = ['A', 'B', 'C', 'D', 'E', 'F',
                        'G', 'H', 'I', 'J', 'K', 'L',
                        'M', 'N', 'O', 'P', 'Q', 'R',
                        'S']
            list_width = [11.63, 26.13, 23.25, 12.63, 12.63, 20.38,
                          20.38, 20.38, 51.75, 51.75, 51.75, 12.38,
                          12.38, 12.38, 12.38, 51.13, 51.13, 51.13,
                          21.5]

            # column width 맞추기
            for idx, item in enumerate(list_col):
                sheet.column_dimensions[item].width = list_width[idx]

            # column filter function 넣기
            # sheet.auto_filer.ref = 'A1:S'+str(tot_rows)

            # 각 행의 옵션 입력
            # Set option Style on Cell
            for mCell in sheet["A1:S" + str(tot_rows)]:
                for cell in mCell:
                    # cell.style = date_style
                    if cell.row == 1:
                        cell.fill = self.blue_fill
                        cell.alignment = self.top_alignment
                        cell.font = self.top_font
                    else:
                        cell.alignment = self.general_alignment
                        cell.font = self.value_normal_font

            # view freeze 'B2' cell
            sheet.freeze_panes = sheet["B2"]

        wb.save(save_path)
        self.setPrint('Report Excel File 옵션 setting 완료...')

    # pandas generate df_result to data
    def generate_report(self):

        try:
            # df.append(df2, ignore_index=True)
            list_keys = []
            list_row_count = []
            current_time = self.getCurrent_time()
            file_path = self.report_path + "\\crawl_report("+ current_time[2] +").xlsx"
            self.report_final_path = file_path
            writer = pd.ExcelWriter(file_path)

            for dict_item in self.list_dict_result:
                list_tuple_data = list(dict_item.items())
                key = list_tuple_data[0][0]
                value = list_tuple_data[0][1]
                if len(value) == 0:
                    dict_empty = {}
                    for item in self.list_cols_2:
                        dict_empty[item] = []

                    df_empty = pd.DataFrame(dict_empty, index=None)
                    df_empty.to_excel(writer, sheet_name=key, index=False)
                    list_keys.append(key)
                    list_row_count.append(0)
                else:
                    df_data = pd.DataFrame(value, index=None)
                    dict_rename_cols = {}
                    for i, item in enumerate(self.list_cols):
                        dict_rename_cols[item] = self.list_cols_2[i]
                    df_data.rename(columns=dict_rename_cols, inplace=True)
                    df_data.to_excel(writer, sheet_name=key, index=False)
                    list_keys.append(key)
                    row_count = df_data.shape[0]
                    list_row_count.append(row_count)

            # 크롤링 데이터 저장
            writer.save()
            self.setPrint('Web 모니터링 결과 Excel File 저장완료...')
            writer.close()
            # 옵션 조절
            self.set_excelOp(list_keys, list_row_count, file_path)
            return True
        except:
            self.setPrint('결과 파일 저장 실패...')
            self.setPrint('Error Occurred : {}. {}, line: {}'.format(sys.exc_info()[0], sys.exc_info()[1], sys.exc_info()[2].tb_lineno))
            return False

if __name__ =='__main__':
    pass
