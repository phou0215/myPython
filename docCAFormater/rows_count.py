import openpyxl
import re
from openpyxl.styles import Alignment, Font, NamedStyle, PatternFill
from openpyxl import formatting, styles, Workbook
from openpyxl.styles.borders import Border, Side

wb_output = openpyxl.load_workbook('C:\\Users\\HANRIM\\Desktop\\uecapa_SM-N981N(TIC_THF).xlsx', data_only=True)
# test_sheet = wb_output['기본단말_Spec정보_설명']

# max_rows = test_sheet.max_row
# max_col = test_sheet.max_column
# print(max_rows)
# print(max_col)
#

def check_condition(ws):
    return_ws = ws
    max_row = ws.max_row
    print('Check Empty Cell Count max_rows : {}'.format(max_row))
    for item in ['A', 'B', 'C']:
        for cell in return_ws[item]:
            if cell.value is None or cell.value == '' or cell.value.lower() in ['n/a', 'na', 'nt', 'n/t']:
                cell.value = '-'
    print('Check Empty Cell job is Completed')
    return return_ws


def extract_condition(ws):
    list_return = []
    for item in ['A', 'B', 'C']:
        list_temp = []
        for cell in ws[item]:
            list_temp.append(cell.value)
        list_return.append(list_temp)
    return list_return

# print(test_sheet['A3'].row, test_sheet['A3'].column,  test_sheet['A3'].coordinate, test_sheet['A3'].value)

list_input_rows = []
ws_lte_cap = wb_output['LTE UE Capabiility']
ws_endc_cap = wb_output['ENDC UE Capability']
ws_lte_ueCap = wb_output['ueCapaInfo_LTE']
ws_5g_nr = wb_output['ueCapaInfo_5G-NR']
ws_spec = wb_output['기본단말_Spec정보']
ws_extract = wb_output['추출정보']
ws_spec_re = check_condition(ws_spec)
ws_extract_re = check_condition(ws_extract)
# A, B, C column 조건 값 추출
a_condi, b_condi, c_condi = extract_condition(ws_extract_re)

# check row index match condition(입력해야 할 대상에 row 주소값 확인)
for idx_2 in range(1, ws_spec_re.max_row + 1):
    value_a = ws_spec_re.cell(row=idx_2, column=1).value
    value_b = ws_spec_re.cell(row=idx_2, column=2).value
    value_c = ws_spec_re.cell(row=idx_2, column=3).value
    # generate condition
    for idx_3, data in enumerate(a_condi):
        if value_a == data and value_b == b_condi[idx_3] and value_c == c_condi[idx_3]:
            list_input_rows.append(idx_2)
            break

list_return = ['', '', '']
list_merged = ws_lte_cap.merged_cells.ranges
list_merged = [str(x) for x in list_merged]
print(list_merged)

# UE category, DL category, UL category value
for cell in ws_lte_cap['A']:

    if cell.value is None:
        continue

    if 'UE Category' in cell.value:

        values = []
        cell_address = cell.coordinate
        cell_rows = [int(cell.row)]
        for item in list_merged:
            if cell_address in item:
                to_row = re.findall('[1-9]{1,2}', item.split(':')[1])
                for idx in range(int(cell.row) + 1, int(to_row[0]) + 1):
                    cell_rows.append(idx)
                break
        # 조회된 row에 있는 값을 가지고 values list에 입력
        print(cell_rows)
        for idx_row in cell_rows:
            c_value = ws_lte_cap['C'+str(idx_row)].value
            e_value = ws_lte_cap['E'+str(idx_row)].value
            if c_value not in [None, '-', '/', '']:
                c_value = re.sub('[A-Za-z]', '', c_value)
                values.append(int(c_value))
            if e_value not in [None, '-', '/', '']:
                e_value = re.sub('[A-Za-z]', '', e_value)
                values.append(int(e_value))
            if c_value in [None, '-', '/', '']:
                values.append(0)
            if e_value in [None, '-', '/', '']:
                values.append(0)
        # top_value
        top_value = max(values)
        list_return[0] = top_value
        continue

    elif 'DL Category' in cell.value:
        values = []
        cell_address = cell.coordinate
        cell_rows = [int(cell.row)]
        for item in list_merged:
            if cell_address in item:
                to_row = re.findall('[1-9]{1,2}', item.split(':')[1])
                for idx in range(int(cell.row) + 1, int(to_row[0]) + 1):
                    cell_rows.append(idx)
                break
        # 조회된 row에 있는 값을 가지고 values list에 입력
        print(cell_rows)
        for idx_row in cell_rows:
            c_value = ws_lte_cap['C'+str(idx_row)].value
            e_value = ws_lte_cap['E'+str(idx_row)].value
            if c_value not in [None, '-', '/', '']:
                c_value = re.sub('[A-Za-z]', '', c_value)
                values.append(int(c_value))
            if e_value not in [None, '-', '/', '']:
                e_value = re.sub('[A-Za-z]', '', e_value)
                values.append(int(e_value))
            if c_value in [None, '-', '/', '']:
                values.append(0)
            if e_value in [None, '-', '/', '']:
                values.append(0)
        # top_value
        top_value = max(values)
        list_return[1] = top_value
        continue

    elif 'UL Category' in cell.value:
        values = []
        cell_address = cell.coordinate
        cell_rows = [int(cell.row)]
        for item in list_merged:
            if cell_address in item:
                to_row = re.findall('[1-9]{1,2}', item.split(':')[1])
                for idx in range(int(cell.row) + 1, int(to_row[0]) + 1):
                    cell_rows.append(idx)
                break
        # 조회된 row에 있는 값을 가지고 values list에 입력
        print(cell_rows)
        for idx_row in cell_rows:
            c_value = ws_lte_cap['C'+str(idx_row)].value
            e_value = ws_lte_cap['E'+str(idx_row)].value
            if c_value not in [None, '-', '/', '']:
                c_value = re.sub('[A-Za-z]', '', c_value)
                values.append(int(c_value))
            if e_value not in [None, '-', '/', '']:
                e_value = re.sub('[A-Za-z]', '', e_value)
                values.append(int(e_value))
            if c_value in [None, '-', '/', '']:
                values.append(0)
            if e_value in [None, '-', '/', '']:
                values.append(0)
        # top_value
        top_value = max(values)
        list_return[2] = top_value
        continue
print(list_return)










# for cell in ws_lte_cap['A']:
#     if cell.value == 'UE Category':
#         print(cell.merged_cells.ranges)


# print('list_index : {}'.format(len(list_input_rows)))
# print(list_input_rows)
# if 'None' in list_input_rows:
#     print('None is in list')
# else:
#     print('All index extract')