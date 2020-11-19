# -*- coding: utf-8 -*-
"""
Created on Thr Oct 18 14:25:12 2019
@author: TestEnC hanrim lee

"""
# 5G와 LTE 구분이 이루어져야 함
# 구분 방법은 Input File Tab에 5G 경우 'ENDC UE Capability' tab이 존재 함
# 기본단말_Spec정보


import os
import sys
import re
import openpyxl
from os.path import expanduser
import threading

#from konlpy.tag import Komoran
from time import sleep
from datetime import datetime

#import pytagcloud
from PyQt5.QtCore import QThread, pyqtSignal
#selenium library
from openpyxl.styles import Alignment, Font, NamedStyle, PatternFill
from openpyxl import formatting, styles, Workbook
from openpyxl.styles.borders import Border, Side
# to print the maximum number of occupied rows in console
# ws.max_row
# to print the maximum number of occupied columns in console
# ws.max_column
class Formater(QThread):

    print_flag = pyqtSignal(str)
    end_flag = pyqtSignal()
    fileCheck_flag = pyqtSignal()
    progress_flag = pyqtSignal()
    count_flag = pyqtSignal()
    dict_result = None
    tot_count = 0

    def __init__(self, filePath, parent=None):
        QThread.__init__(self, parent)

        self.file_names = filePath
        self.list_files = self.file_names.split(",")
        self.list_out_files = []
        self.dict_out = {}
        self.dict_readData = {}
        # self.list_sheet_names = []

        self.home = expanduser("~")

        self.end_count = "n"
        self.totalRows = 0
        self.currentRow = 0
        self.current_path = os.getcwd()

        # style fill pattern
        # FF0000 red
        # 0000FF blue

        #Pattern_Fill
        self.gray_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        self.dark_gray_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

        # style font color and size
        self.red_font = Font(name='맑은 고딕', size=10, bold=True, color='FF0000')
        self.blue_font = Font(name='맑은 고딕', size=10, bold=True, color='0000FF')

        # style Alignment
        self.general_alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        self.top_alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")

        # # style border
        self.thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))


    # 분석 처리 개수 체크 함수
    def getCountRows(self):

        while True:
            if self.end_count is "n":
                sleep(0.5)
                self.count_flag.emit()
            else:
                break

    # 로그 문자 처리 함수
    def setPrintText(self, text):

        strToday = datetime.today().strftime("%Y-%m-%d %H:%M:%S")
        text = self.find_between(text, "/s", "/e")
        print_text = strToday+":\n"+text+"\n"
        self.print_flag.emit("{}".format(print_text))

    # 쓰레드 종료 함수
    def stop(self):
        sleep(0.5)
        self.terminate()

    # 특수 문자 제거 함수
    def removeString(self, text):

        tempText = re.sub('[-=+,#/\?^$@*\"※~&%ㆍ!』\\‘|\(\)\[\]\<\>\{\}`><]\'', '', text)
        return tempText

    # 문장 앞 부터 조건에 맞는 문자열 substring
    def find_between(self, s, first, last):
        try:
            returnData = ""
            start = s.index(first)+len(first)
            end = s.index(last, start)
            returnData = s[start:end]
            return returnData
        except ValueError:
            return returnData

    # 문장 뒤 부터 조건에 맞는 문자열 substring
    def find_between_r(self, s, first, last ):
        try:
            returnData = ""
            start = s.rindex(first)+len(first)
            end = s.rindex(last, start)
            returnData = s[start:end]
            return returnData
        except ValueError:
            return returnData

    # float num check a point
    def check_num(self, num):

        return_data = None
        if num is None or num == '':
            return_data = '-'
        else:
            try:
                return_data = '%.2f'%float(num)
            except:
                return_data = str(num)

        return return_data

    # check calculate comparison
    def cal_comparison(self, standard, measure):

        return_data = None
        try:
            return_data = self.check_num(abs(round(abs(float(measure)) - float(standard), 2)))
        except:
            return_data = '-'

        return return_data

    # check convert num available
    def isNumber(self, string_data):

        try:
            temp_data = float(string_data)
            return True
        except:
            return False

    def count_duplicate(self, list_data, element):
        return_count = 0
        for data in list_data:
            if element in data:
                return_count += 1
        return return_count


    ####################################################################__condition method group__####################################################################
    # LTE UE Capabiility Tab UE/UL/DL category (3 unit)
    def get_conditional_1(self, ws, index):

        list_return = ['X', 'X', 'X']
        list_merged = ws.merged_cells.ranges
        list_merged = [str(x) for x in list_merged]
        # UE category, DL category, UL category value

        for cell in ws['A']:
            if cell.value is None:
                continue
            if 'ue category' in str(cell.value).lower().strip():

                values = []
                cell_address = cell.coordinate
                cell_rows = [int(cell.row)]
                for item in list_merged:
                    if cell_address in item:
                        to_row = re.findall('[1-9]{1,2}', item.split(':')[1])
                        for idx in range(int(cell.row) + 1, int(to_row[0]) + 1):
                            cell_rows.append(idx)
                        break
                # 조회된 row에 있는 값을 가지고 values list에 입력
                for idx_row in cell_rows:
                    c_value = ws['C' + str(idx_row)].value
                    e_value = ws['E' + str(idx_row)].value
                    if c_value not in [None, '-', '/', '']:
                        c_value = re.sub('[A-Za-z]', '', c_value)
                        values.append(int(c_value))
                    if e_value not in [None, '-', '/', '']:
                        e_value = re.sub('[A-Za-z]', '', e_value)
                        values.append(int(e_value))
                    if c_value in [None, '-', '/', '']:
                        values.append(0)
                    if e_value in [None, '-', '/', '']:
                        values.append(0)
                # top_value
                top_value = max(values)
                if top_value != 0:
                    list_return[0] = top_value
                continue
            elif 'dl category' in str(cell.value).lower().strip():
                values = []
                cell_address = cell.coordinate
                cell_rows = [int(cell.row)]
                for item in list_merged:
                    if cell_address in item:
                        to_row = re.findall('[1-9]{1,2}', item.split(':')[1])
                        for idx in range(int(cell.row) + 1, int(to_row[0]) + 1):
                            cell_rows.append(idx)
                        break
                # 조회된 row에 있는 값을 가지고 values list에 입력
                for idx_row in cell_rows:
                    c_value = ws['C' + str(idx_row)].value
                    e_value = ws['E' + str(idx_row)].value
                    if c_value not in [None, '-', '/', '']:
                        c_value = re.sub('[A-Za-z]', '', c_value)
                        values.append(int(c_value))
                    if e_value not in [None, '-', '/', '']:
                        e_value = re.sub('[A-Za-z]', '', e_value)
                        values.append(int(e_value))
                    if c_value in [None, '-', '/', '']:
                        values.append(0)
                    if e_value in [None, '-', '/', '']:
                        values.append(0)
                # top_value
                top_value = max(values)
                if top_value != 0:
                    list_return[1] = top_value
                continue
            elif 'ul category' in str(cell.value).lower().strip():
                values = []
                cell_address = cell.coordinate
                cell_rows = [int(cell.row)]
                for item in list_merged:
                    if cell_address in item:
                        to_row = re.findall('[1-9]{1,2}', item.split(':')[1])
                        for idx in range(int(cell.row) + 1, int(to_row[0]) + 1):
                            cell_rows.append(idx)
                        break
                # 조회된 row에 있는 값을 가지고 values list에 입력
                for idx_row in cell_rows:
                    c_value = ws['C' + str(idx_row)].value
                    e_value = ws['E' + str(idx_row)].value
                    if c_value not in [None, '-', '/', '']:
                        c_value = re.sub('[A-Za-z]', '', c_value)
                        values.append(int(c_value))
                    if e_value not in [None, '-', '/', '']:
                        e_value = re.sub('[A-Za-z]', '', e_value)
                        values.append(int(e_value))
                    if c_value in [None, '-', '/', '']:
                        values.append(0)
                    if e_value in [None, '-', '/', '']:
                        values.append(0)
                # top_value
                top_value = max(values)
                if top_value != 0:
                    list_return[2] = top_value
                continue
        self.setPrintText('/s {} Method 1 Return Array : {} /e'.format(index, list_return))
        return list_return

    # LTE UE Capabiility Tab DL256QAM/ UL64QAM / ULCA / MC-PUSCH / 4X4 MIMO / CA 지원여부 / Max CC(7 unit)
    # [None, '-', '/', '']
    def get_conditional_2(self, ws, index):
        list_return = ['N', 'N', 'N', 'N', 'N', 'N', 'N']
        #####__DL256QAM Data Get__#####
        for cell in ws['B']:
            if cell.value is None:
                continue
            if 'dl modulation' in str(cell.value).lower().strip():
                cell_row = cell.row
                c_value = ws['C'+str(cell_row)].value
                if c_value is None:
                    list_return[0] = 'N'
                elif '256qam' in c_value.lower().strip():
                    list_return[0] = 'Y'
                else:
                    list_return[0] = 'N'
                break
        #####__UL64QAM Data Get__#####
        for cell in ws['D']:
            if cell.value is None:
                continue
            if 'ul modulation' in str(cell.value).lower().strip():
                cell_row = cell.row
                c_value = ws['E'+str(cell_row)].value
                if c_value is None:
                    list_return[1] = 'N'
                elif '64qam' in c_value.lower().strip():
                    list_return[1] = 'Y'
                else:
                    list_return[1] = 'N'
                break
        #####__ULCA Data Get__#####
        for cell in ws['D']:
            if cell.value is None:
                continue
            if 'uplink' in str(cell.value).lower().strip():
                cell_row = cell.row
                max_row = ws.max_row
                flag_support = False
                for idx in range(cell_row + 1, max_row + 1):
                    ca_value = ws['D' + str(idx)].value
                    if '-' in ca_value:
                        flag_support = True
                        break
                if flag_support:
                    list_return[2] = 'Y'
                else:
                    list_return[2] = 'N'
                break
        #####__MC-PUSCH Data Get__#####
        for cell in ws['D']:
            if cell.value is None:
                continue
            if 'mc-pusch' in str(cell.value).lower().strip():
                cell_row = cell.row
                e_value = ws['E'+str(cell_row)].value
                if e_value is None:
                    list_return[3] = 'N'
                if e_value.lower().strip() == 'supported':
                    list_return[3] = 'Y'
                else:
                    list_return[3] = 'N'
                break
        #####__4X4 MIMO Data Get__#####
        for cell in ws['E']:
            if cell.value is None:
                continue
            if 'mimo layer' in str(cell.value).lower().strip():
                cell_row = cell.row
                max_row = ws.max_row
                flag_support = False
                for idx in range(cell_row + 1, max_row + 1):
                    ca_value = ws['E' + str(idx)].value
                    if ca_value not in [None, '-', '/', '']:
                        flag_support = True
                        break
                if flag_support:
                    list_return[4] = 'Y'
                else:
                    list_return[4] = 'N'
                break
        #####__CA 지원여부 Data Get__#####
        for cell in ws['B']:
            if cell.value is None:
                continue
            if 'combination' in str(cell.value).lower().strip():
                cell_row = cell.row
                max_row = ws.max_row
                flag_support = False
                for idx in range(cell_row + 1, max_row + 1):
                    ca_value = ws['B' + str(idx)].value
                    if '+' in ca_value:
                        flag_support = True
                        break
                if flag_support:
                    list_return[5] = 'Y'
                else:
                    list_return[5] = 'N'
                break
        #####__MAX CC Data Get__#####
        for cell in ws['B']:
            if cell.value is None:
                continue
            if 'combination' in str(cell.value).lower().strip():
                cell_row = cell.row
                max_row = ws.max_row
                max_cc = 0
                for idx in range(cell_row + 1, max_row + 1):
                    ca_value = ws['B' + str(idx)].value
                    if '+' in ca_value:
                        list_band = ca_value.split('+')
                        if len(list_band) > max_cc:
                            max_cc = len(list_band)

                if max_cc > 0:
                    list_return[6] = max_cc
                else:
                    list_return[6] = 'N'
                break
        # retrun values
        self.setPrintText('/s {} Method 2 Return Array : {} /e'.format(index, list_return))
        return list_return

    # ueCapaInfo Tab == > CRDX(Short) / CRDX(Long) / L2W H/O / SRVCC(B2) / SRVCC(CS H/O) / ANR(Inter Freq.)
    # / ANR(Inter RAT) / TTIB / MFBI / A6 Event / RoHC(Profile) 1 / RoHC(Profile) 2/ RoHC(Profile) 4(13 unit)
    def get_conditional_3(self, ws, index):
        list_return = ['', '', '', '', '', '', '', '', '', '', '', '', '']
        tot_job_count = 0
        ##__ueCapaInfo data extract__##
        for cell in ws['A']:
            if cell.value is None:
                continue
            if tot_job_count == 13:
                break
            # get just one value extract
            if str(cell.value).lower().strip() == 'supportedrohc-profiles':
                cell_row = cell.row
                max_row = ws.max_row
                supported_job_count = 0

                for idx in range(cell_row + 1, max_row + 1):
                    ca_value = str(ws['A' + str(idx)].value).lower().strip()
                    if supported_job_count == 3:
                        break
                    #####__profile0x0001__#####
                    if 'profile0x0001-r15' in ca_value:
                        if 'true' in ca_value:
                            list_return[10] = 'Y'
                        else:
                            list_return[10] = 'N'
                        supported_job_count = supported_job_count + 1
                        tot_job_count = tot_job_count + 1
                    #####__profile0x0002__#####
                    if 'profile0x0002-r15' in ca_value:
                        if 'true' in ca_value:
                            list_return[11] = 'Y'
                        else:
                            list_return[11] = 'N'
                        supported_job_count = supported_job_count + 1
                        tot_job_count = tot_job_count + 1
                    #####__profile0x0004__#####
                    if 'profile0x0004-r15' in ca_value:
                        if 'true' in ca_value:
                            list_return[12] = 'Y'
                        else:
                            list_return[12] = 'N'
                        supported_job_count = supported_job_count + 1
                        tot_job_count = tot_job_count + 1

            # get 1 ~ 10 unit extract
            if 'feature group indicators' in str(cell.value).lower().strip():
                cell_row = cell.row
                max_row = ws.max_row
                feature_job_count = 0
                cs_ho_values = ['', '', '']
                for idx in range(cell_row + 1, max_row + 1):
                    ca_value = str(ws['A' + str(idx)].value).lower().strip()
                    if feature_job_count == 12:
                        break
                    #####__CRDX(Short)__#####
                    if 'pc_featrgrp_4 :' in ca_value:
                        list_info = ca_value.split(':')
                        if list_info[1].lower().strip() == 'support':
                            list_return[0] = 'Y'
                        else:
                            list_return[0] = 'N'
                        tot_job_count = tot_job_count + 1
                        feature_job_count = feature_job_count + 1
                    #####__CRDX(Long)__#####
                    if 'pc_featrgrp_5 :' in ca_value:
                        list_info = ca_value.split(':')
                        if list_info[1].lower().strip() == 'support':
                            list_return[1] = 'Y'
                        else:
                            list_return[1] = 'N'
                        tot_job_count = tot_job_count + 1
                        feature_job_count = feature_job_count + 1
                    #####__L2W H/O__#####
                    if 'pc_featrgrp_8 :' in ca_value:
                        list_info = ca_value.split(':')
                        if list_info[1].lower().strip() == 'support':
                            list_return[2] = 'Y'
                        else:
                            list_return[2] = 'N'
                        tot_job_count = tot_job_count + 1
                        feature_job_count = feature_job_count + 1
                    #####__SRVCC(B2)__#####
                    if 'pc_featrgrp_22 :' in ca_value:
                        list_info = ca_value.split(':')
                        if list_info[1].lower().strip() == 'support':
                            list_return[3] = 'Y'
                        else:
                            list_return[3] = 'N'
                        tot_job_count = tot_job_count + 1
                        feature_job_count = feature_job_count + 1
                    #####__featrgrp_13__#####
                    if 'pc_featrgrp_13 :' in ca_value:
                        list_info = ca_value.split(':')
                        cs_ho_values[0] = list_info[1].lower().strip()
                        feature_job_count = feature_job_count + 1
                    #####__featrgrp_25__#####
                    if 'pc_featrgrp_25 :' in ca_value:
                        list_info = ca_value.split(':')
                        cs_ho_values[1] = list_info[1].lower().strip()
                        feature_job_count = feature_job_count + 1
                    #####__featrgrp_27__#####
                    if 'pc_featrgrp_27 :' in ca_value:
                        list_info = ca_value.split(':')
                        cs_ho_values[2] = list_info[1].lower().strip()
                        feature_job_count = feature_job_count + 1
                    #####__ANR(Inter Freq.)__#####
                    if 'pc_featrgrp_18 :' in ca_value:
                        list_info = ca_value.split(':')
                        if list_info[1].lower().strip() == 'support':
                            list_return[5] = 'Y'
                        else:
                            list_return[5] = 'N'
                        tot_job_count = tot_job_count + 1
                        feature_job_count = feature_job_count + 1
                    #####__ANR(Inter RAT)__#####
                    if 'pc_featrgrp_19 :' in ca_value:
                        list_info = ca_value.split(':')
                        if list_info[1].lower().strip() == 'support':
                            list_return[6] = 'Y'
                        else:
                            list_return[6] = 'N'
                        tot_job_count = tot_job_count + 1
                        feature_job_count = feature_job_count + 1
                    #####__TTIB__#####
                    if 'pc_featrgrp_28 :' in ca_value:
                        list_info = ca_value.split(':')
                        if list_info[1].lower().strip() == 'support':
                            list_return[7] = 'Y'
                        else:
                            list_return[7] = 'N'
                        tot_job_count = tot_job_count + 1
                        feature_job_count = feature_job_count + 1
                    #####__MFBI__#####
                    if 'pc_featrgrp_31 :' in ca_value:
                        list_info = ca_value.split(':')
                        if list_info[1].lower().strip() == 'support':
                            list_return[8] = 'Y'
                        else:
                            list_return[8] = 'N'
                        tot_job_count = tot_job_count + 1
                        feature_job_count = feature_job_count + 1
                    #####__A6 Event__#####
                    if 'pc_featrgrp_111 :' in ca_value:
                        list_info = ca_value.split(':')
                        if list_info[1].lower().strip() == 'support':
                            list_return[9] = 'Y'
                        else:
                            list_return[9] = 'N'
                        tot_job_count = tot_job_count + 1
                        feature_job_count = feature_job_count + 1

                #####__SRVCC(CS H/O)__#####
                flag_cs = True
                for item in cs_ho_values:
                    if item == 'support':
                        pass
                    else:
                        flag_cs = False
                        break
                if flag_cs:
                    list_return[4] = 'Y'
                else:
                    list_return[4] = 'N'
                tot_job_count = tot_job_count + 1

        self.setPrintText('/s {} Method 3 Return Array : {} /e'.format(index, list_return))
        return list_return

    # LTE UE Capabiility tab 지원주파수 EUTRA-FDD(B1) / EUTRA-FDD(B3) /EUTRA-FDD(B5) / EUTRA-FDD(B7) / EUTRA-FDD(B8) (5 unit)
    def get_conditional_4(self, ws, index):

        list_return = ['N', 'N', 'N', 'N', 'N']
        for cell in ws['A']:
            if cell.value is None:
                continue
            if str(cell.value).lower().strip() == 'single band':
                row_idx = cell.row
                c_value = str(ws['C' + str(row_idx)].value).lower().strip()
                # band 1 check
                if '1' in c_value:
                    list_return[0] = 'Y'
                # band 3 check
                if '3' in c_value:
                    list_return[1] = 'Y'
                # band 5 check
                if '5' in c_value:
                    list_return[2] = 'Y'
                # band 7 check
                if '7' in c_value:
                    list_return[3] = 'Y'
                # band 8 check
                if '8' in c_value:
                    list_return[4] = 'Y'
                break
        # return band info
        self.setPrintText('/s {} Method 4 Return Array : {} /e'.format(index, list_return))
        return list_return

    #  LTE UE Capabiility tab 지원 CA ==> 5CA(2UL 지원여부) / 5CA(B1+B3+B5+B7+B7) / 4CA(2UL 지원여부) / 4CA(B1+B3+B5+B7)
    #  / 3DL+1UL(B1+B3+B5) / 3DL+2UL(B1+B3+B5) / 2DL+1UL(B1+B3) / 2DL+2UL(B1+B3) (35 unit)
    def get_conditional_5(self, ws, index):

        list_return = []
        # set default value list_return
        for i in range(35):
            list_return.append('N')
        # variable of band and ca info
        list_band_info = []
        list_ca_info = []
        # extract start
        for cell in ws['B']:
            if cell.value is None:
                continue
            # first find 'band combination'
            if str(cell.value).lower().strip() == 'band combination':
                cell_row = cell.row
                max_row = ws.max_row
                # append band info in list_band_info
                for idx in range(cell_row + 1, max_row + 1):
                    band_value = str(ws['B'+str(idx)].value).lower().strip()
                    ca_value = str(ws['C'+str(idx)].value).lower().strip()
                    ca_value = ca_value.replace('a', '')
                    list_band_info.append(band_value)
                    list_ca_info.append(ca_value)
                # check band info
                for idx, item in enumerate(list_band_info):
                    # size CA count
                    list_ca_band = item.split('+')
                    ca_count = len(list_ca_band)
                    # check uplink support 2UL in 5CA
                    if ca_count == 5:
                        ul_count = 0
                        # check ul count on ca band
                        for ca in list_ca_band:
                            if 'ul' in ca:
                                ul_count = ul_count + 1
                        if ul_count >= 2:
                            list_return[0] = 'Y'

                    # check 1+3+5+7+7 band combination  in 5CA
                    if ca_count == 5:
                        ca_info = list_ca_info[idx]
                        list_ca = ca_info.split('-')
                        band_1_count = list_ca.count('1')
                        band_3_count = list_ca.count('3')
                        band_5_count = list_ca.count('5')
                        band_7_count = list_ca.count('7')
                        if band_1_count == 1 and band_3_count == 1 and band_5_count == 1 and band_7_count == 2:
                            list_return[1] = 'Y'

                    # check uplink support 2UL in 4CA
                    if ca_count == 4:
                        ul_count = 0
                        # check ul count on ca band
                        for ca in list_ca_band:
                            if 'ul' in ca:
                                ul_count = ul_count + 1
                        if ul_count >= 2:
                            list_return[2] = 'Y'

                    # check 1+3+5+7/1+3+7+7/1+5+7+7/3+5+7+7 band combination  in 4CA
                    if ca_count == 4:
                        ca_info = list_ca_info[idx]
                        list_ca = ca_info.split('-')
                        band_1_count = list_ca.count('1')
                        band_3_count = list_ca.count('3')
                        band_5_count = list_ca.count('5')
                        band_7_count = list_ca.count('7')
                        if band_1_count == 1 and band_3_count == 1 and band_5_count == 1 and band_7_count == 1:
                            list_return[3] = 'Y'
                        if band_1_count == 1 and band_3_count == 1 and band_7_count == 2:
                            list_return[4] = 'Y'
                        if band_1_count == 1 and band_5_count == 1 and band_7_count == 2:
                            list_return[5] = 'Y'
                        if band_3_count == 1 and band_5_count == 1 and band_7_count == 2:
                            list_return[6] = 'Y'

                    # Check Band 3CA
                    if ca_count == 3:
                        ca_info = list_ca_info[idx]
                        list_ca = ca_info.split('-')
                        band_1_count = list_ca.count('1')
                        band_3_count = list_ca.count('3')
                        band_5_count = list_ca.count('5')
                        band_7_count = list_ca.count('7')

                        # check 1+3+5 band combination
                        if band_1_count == 1 and band_3_count == 1 and band_5_count == 1:
                            list_return[7] = 'Y'
                            ul_count = 0
                            # check ul count on ca band
                            for ca in list_ca_band:
                                if 'ul' in ca:
                                    ul_count = ul_count + 1
                            if ul_count >= 2:
                                list_return[14] = 'Y'

                        # check 1+3+7 band combination
                        elif band_1_count == 1 and band_3_count == 1 and band_7_count == 1:
                            list_return[8] = 'Y'
                            ul_count = 0
                            # check ul count on ca band
                            for ca in list_ca_band:
                                if 'ul' in ca:
                                    ul_count = ul_count + 1
                            if ul_count >= 2:
                                list_return[15] = 'Y'

                        # check 1+5+7 band combination
                        elif band_1_count == 1 and band_5_count == 1 and band_7_count == 1:
                            list_return[9] = 'Y'
                            ul_count = 0
                            # check ul count on ca band
                            for ca in list_ca_band:
                                if 'ul' in ca:
                                    ul_count = ul_count + 1
                            if ul_count >= 2:
                                list_return[16] = 'Y'

                        # check 3+5+7 band combination
                        elif band_3_count == 1 and band_5_count == 1 and band_7_count == 1:
                            list_return[10] = 'Y'
                            ul_count = 0
                            # check ul count on ca band
                            for ca in list_ca_band:
                                if 'ul' in ca:
                                    ul_count = ul_count + 1
                            if ul_count >= 2:
                                list_return[17] = 'Y'

                        # check 1+7+7 band combination
                        elif band_1_count == 1 and band_7_count == 2:
                            list_return[11] = 'Y'
                            ul_count = 0
                            # check ul count on ca band
                            for ca in list_ca_band:
                                if 'ul' in ca:
                                    ul_count = ul_count + 1
                            if ul_count >= 2:
                                list_return[18] = 'Y'

                        # check 3+7+7 band combination
                        elif band_3_count == 1 and band_7_count == 2:
                            list_return[12] = 'Y'
                            ul_count = 0
                            # check ul count on ca band
                            for ca in list_ca_band:
                                if 'ul' in ca:
                                    ul_count = ul_count + 1
                            if ul_count >= 2:
                                list_return[19] = 'Y'

                        # check 5+7+7 band combination
                        elif band_5_count == 1 and band_7_count == 2:
                            list_return[13] = 'Y'
                            ul_count = 0
                            # check ul count on ca band
                            for ca in list_ca_band:
                                if 'ul' in ca:
                                    ul_count = ul_count + 1
                            if ul_count >= 2:
                                list_return[20] = 'Y'

                    # Check Band 2CA
                    if ca_count == 2:
                        ca_info = list_ca_info[idx]
                        list_ca = ca_info.split('-')
                        band_1_count = list_ca.count('1')
                        band_3_count = list_ca.count('3')
                        band_5_count = list_ca.count('5')
                        band_7_count = list_ca.count('7')

                        # check 1+3 band combination
                        if band_1_count == 1 and band_3_count == 1:
                            list_return[21] = 'Y'
                            ul_count = 0
                            # check ul count on ca band
                            for ca in list_ca_band:
                                if 'ul' in ca:
                                    ul_count = ul_count + 1
                            if ul_count >= 2:
                                list_return[28] = 'Y'

                        # check 1+5 band combination
                        elif band_1_count == 1 and band_5_count == 1:
                            list_return[22] = 'Y'
                            ul_count = 0
                            # check ul count on ca band
                            for ca in list_ca_band:
                                if 'ul' in ca:
                                    ul_count = ul_count + 1
                            if ul_count >= 2:
                                list_return[29] = 'Y'

                        # check 1+7 band combination
                        elif band_1_count == 1 and band_7_count == 1:
                            list_return[23] = 'Y'
                            ul_count = 0
                            # check ul count on ca band
                            for ca in list_ca_band:
                                if 'ul' in ca:
                                    ul_count = ul_count + 1
                            if ul_count >= 2:
                                list_return[30] = 'Y'

                        # check 3+5 band combination
                        elif band_3_count == 1 and band_5_count == 1:
                            list_return[24] = 'Y'
                            ul_count = 0
                            # check ul count on ca band
                            for ca in list_ca_band:
                                if 'ul' in ca:
                                    ul_count = ul_count + 1
                            if ul_count >= 2:
                                list_return[31] = 'Y'

                        # check 3+7 band combination
                        elif band_3_count == 1 and band_7_count == 1:
                            list_return[25] = 'Y'
                            ul_count = 0
                            # check ul count on ca band
                            for ca in list_ca_band:
                                if 'ul' in ca:
                                    ul_count = ul_count + 1
                            if ul_count >= 2:
                                list_return[32] = 'Y'

                        # check 5+7 band combination
                        elif band_5_count == 1 and band_7_count == 1:
                            list_return[26] = 'Y'
                            ul_count = 0
                            # check ul count on ca band
                            for ca in list_ca_band:
                                if 'ul' in ca:
                                    ul_count = ul_count + 1
                            if ul_count >= 2:
                                list_return[33] = 'Y'

                        # check 7+7 band combination
                        elif band_7_count == 2:
                            list_return[27] = 'Y'
                            ul_count = 0
                            # check ul count on ca band
                            for ca in list_ca_band:
                                if 'ul' in ca:
                                    ul_count = ul_count + 1
                            if ul_count >= 2:
                                list_return[34] = 'Y'
                # stop method
                break
        # return band info
        self.setPrintText('/s {} Method 5 Return Array : {} /e'.format(index, list_return))
        return list_return

    #  LTE UE Capabiility tab 지원 MIMO ==>(42 unit)
    def get_conditional_6(self, ws, index):

        list_return = []
        # set default value list_return
        for i in range(42):
            list_return.append('N')
        # variable of band and ca info
        list_mimo_info = []
        list_band_info = []
        # extract start
        for cell in ws['E']:
            if cell.value is None:
                continue
            # first find 'band combination'
            if '4x4 mimo layer' in str(cell.value).lower().strip():
                cell_row = cell.row
                max_row = ws.max_row
                # append MIMO and band info in list
                for idx in range(cell_row + 1, max_row + 1):
                    mimo_value = '-'
                    band_value = '-'
                    if ws['E'+str(idx)].value not in [None, '-', '']:
                        mimo_value = str(ws['E'+str(idx)].value).lower().replace(' ', '')
                    if ws['B'+str(idx)].value not in [None, '-', '']:
                        band_value = str(ws['B'+str(idx)].value).lower()
                    list_mimo_info.append(mimo_value)
                    list_band_info.append(band_value)

                # check band info
                for idx, item in enumerate(list_band_info):

                    # size CA count
                    list_ca_band = item.split('+')
                    ca_count = len(list_ca_band)
                    mimo_value = list_mimo_info[idx]
                    list_mimo = mimo_value.split(',')

                    # MIMO b1 포함 CA 조합 조회
                    if 'b1' in mimo_value:
                        if ca_count == 1:
                            list_return[0] = 'Y'
                        elif ca_count == 2:
                            list_return[1] = 'Y'
                        elif ca_count == 3:
                            list_return[2] = 'Y'
                        elif ca_count == 4:
                            list_return[3] = 'Y'
                        elif ca_count == 5:
                            list_return[4] = 'Y'

                    # MIMO b3 포함 CA 조합 조회
                    if 'b3' in mimo_value:
                        if ca_count == 1:
                            list_return[5] = 'Y'
                        elif ca_count == 2:
                            list_return[6] = 'Y'
                        elif ca_count == 3:
                            list_return[7] = 'Y'
                        elif ca_count == 4:
                            list_return[8] = 'Y'
                        elif ca_count == 5:
                            list_return[9] = 'Y'

                    # MIMO b7 포함 CA 조합 조회
                    if 'b7' in mimo_value:
                        if ca_count == 1:
                            list_return[10] = 'Y'
                        elif ca_count == 2:
                            list_return[11] = 'Y'
                        elif ca_count == 3:
                            list_return[12] = 'Y'
                        elif ca_count == 4:
                            list_return[13] = 'Y'
                        elif ca_count == 5:
                            list_return[14] = 'Y'

                    # MIMO b1 + b3 포함 CA 조합 조회
                    if 'b1' in mimo_value and 'b3' in mimo_value:
                        if ca_count == 2:
                            list_return[15] = 'Y'
                        elif ca_count == 3:
                            list_return[16] = 'Y'
                        elif ca_count == 4:
                            list_return[17] = 'Y'
                        elif ca_count == 5:
                            list_return[18] = 'Y'

                    # MIMO b3 + b7 포함 CA 조합 조회
                    if 'b3' in mimo_value and 'b7' in mimo_value:
                        if ca_count == 2:
                            list_return[19] = 'Y'
                        elif ca_count == 3:
                            list_return[20] = 'Y'
                        elif ca_count == 4:
                            list_return[21] = 'Y'
                        elif ca_count == 5:
                            list_return[22] = 'Y'

                    # MIMO b1 + b7 포함 CA 조합 조회
                    if 'b1' in mimo_value and 'b7' in mimo_value:
                        if ca_count == 2:
                            list_return[23] = 'Y'
                        elif ca_count == 3:
                            list_return[24] = 'Y'
                        elif ca_count == 4:
                            list_return[25] = 'Y'
                        elif ca_count == 5:
                            list_return[26] = 'Y'

                    # MIMO b7 + b7 포함 CA 조합 조회
                    if list_mimo.count("b7") >= 2:
                        if ca_count == 2:
                            list_return[27] = 'Y'
                        elif ca_count == 3:
                            list_return[28] = 'Y'
                        elif ca_count == 4:
                            list_return[29] = 'Y'
                        elif ca_count == 5:
                            list_return[30] = 'Y'

                    # MIMO b1 + b3 + b7 포함 CA 조합 조회
                    if 'b1' in mimo_value and 'b3' in mimo_value and 'b7' in mimo_value:
                        if ca_count == 3:
                            list_return[31] = 'Y'
                        elif ca_count == 4:
                            list_return[32] = 'Y'
                        elif ca_count == 5:
                            list_return[33] = 'Y'

                    # MIMO b1 + b7 + b7 포함 CA 조합 조회
                    if 'b1' in mimo_value and list_mimo.count("b7") == 2:
                        if ca_count == 3:
                            list_return[34] = 'Y'
                        elif ca_count == 4:
                            list_return[35] = 'Y'
                        elif ca_count == 5:
                            list_return[36] = 'Y'

                    # MIMO b3 + b7 + b7 포함 CA 조합 조회
                    if 'b3' in mimo_value and list_mimo.count("b7") == 2:
                        if ca_count == 3:
                            list_return[37] = 'Y'
                        elif ca_count == 4:
                            list_return[38] = 'Y'
                        elif ca_count == 5:
                            list_return[39] = 'Y'

                    # MIMO b1 + b3 + b7 + b7 포함 CA 조합 조회
                    if 'b1' in mimo_value and 'b3' in mimo_value and list_mimo.count("b7") == 2:
                        if ca_count == 4:
                            list_return[40] = 'Y'
                        elif ca_count == 5:
                            list_return[41] = 'Y'

                # stop method
                break


        # return band info
        self.setPrintText('/s {} Method 6 Return Array : {} /e'.format(index, list_return))
        return list_return

    # extract condition in extract sheet 'A', 'B', 'C' column
    def extract_condition(self, ws):

        list_return = []
        for item in ['A', 'B', 'C']:
            list_temp = []
            for cell in ws[item]:
                list_temp.append(cell.value)
            list_return.append(list_temp)
        return list_return

    # check convert num available
    def check_empty(self, string_data):

        return_data = ''
        if string_data is None or string_data == '' or string_data.lower() in ['n/a', 'na', 'nt', 'n/t']:
            return_data = '-'
        else:
            return_data = str(string_data)
        return return_data

    # check empty cell in sheet
    def check_condition(self, ws, sheet_name):
        return_ws = ws
        max_row = ws.max_row
        self.setPrintText('/s {} sheet Check Empty Cell Count max_rows : {}/e'.format(sheet_name, max_row))
        for item in ['A', 'B', 'C']:
            for cell in return_ws[item]:
                if cell.value is None or cell.value == '' or cell.value.lower() in ['n/a', 'na', 'nt', 'n/t']:
                    cell.value = '-'
        self.setPrintText('/s {} sheet Check Empty Cell job is Completed /e'.format(sheet_name))
        return return_ws

    # generate ca file data sheet and save
    def generate_sheet(self, file_in_path, file_out_path, file_idx):

        try:
            list_input_rows = []
            wb_output = openpyxl.load_workbook(file_in_path, data_only=True)
            list_sheets = wb_output.sheetnames
            sheet_type = 1
            # 모든 필요 시트 충족 조건
            if 'LTE UE Capabiility' in list_sheets and ('ueCapaInfo_LTE' in list_sheets or 'ueCapaInfo' in list_sheets)and \
                '추출정보' in list_sheets and '기본단말_Spec정보' in list_sheets:
                # Capabiility sheet type 확인
                if 'ueCapaInfo_LTE' in list_sheets:
                    sheet_type = 2

                ws_lte_cap = wb_output['LTE UE Capabiility']
                # Capabiility '1'
                if sheet_type == 1:
                    ws_lte_ueCap = wb_output['ueCapaInfo']
                # Capabiility '2'
                else:
                    ws_lte_ueCap = wb_output['ueCapaInfo_LTE']

                ws_spec = wb_output['기본단말_Spec정보']
                ws_extract = wb_output['추출정보']

                ws_spec_re = self.check_condition(ws_spec, '기본단말_Spec정보')
                ws_extract_re = self.check_condition(ws_extract, '추출정보')

                # A, B, C column 조건 값 추출
                a_condi, b_condi, c_condi = self.extract_condition(ws_extract_re)

                # check row index match condition(입력해야 할 대상에 row 주소값 확인)
                for idx in range(1, ws_spec_re.max_row + 1):
                    value_a = ws_spec_re.cell(row=idx, column=1).value
                    value_b = ws_spec_re.cell(row=idx, column=2).value
                    value_c = ws_spec_re.cell(row=idx, column=3).value
                    # generate condition
                    for idx_2, data in enumerate(a_condi):
                        if value_a == data and value_b == b_condi[idx_2] and value_c == c_condi[idx_2]:
                            list_input_rows.append(idx)
                            break

                # Check len count
                if len(list_input_rows) != 105:
                    self.setPrintText('/s "{}" 파일의 추출정보에 있는 값 중 "기본단말_Spec정보" 에 없는 것이 있습니다. 정보를 다시 확인해 주세요. /e'.format(file_in_path))
                    return

                ##########################################__1 method start__##########################################
                list_tot_returns = self.get_conditional_1(ws_lte_cap, file_idx)
                self.currentRow = 1
                self.progress_flag.emit()
                ##########################################__2 method start__##########################################
                list_tot_returns.extend(self.get_conditional_2(ws_lte_cap, file_idx))
                self.currentRow = 2
                self.progress_flag.emit()
                ##########################################__3 method start__##########################################
                list_tot_returns.extend(self.get_conditional_3(ws_lte_ueCap, file_idx))
                self.currentRow = 3
                self.progress_flag.emit()
                ##########################################__4 method start__##########################################
                list_tot_returns.extend(self.get_conditional_4(ws_lte_cap, file_idx))
                self.currentRow = 4
                self.progress_flag.emit()
                ##########################################__5 method start__##########################################
                list_tot_returns.extend(self.get_conditional_5(ws_lte_cap, file_idx))
                self.currentRow = 5
                self.progress_flag.emit()
                ##########################################__6 method start__##########################################
                list_tot_returns.extend(self.get_conditional_6(ws_lte_cap, file_idx))
                self.currentRow = 6
                self.progress_flag.emit()
                ##########################################__spec 시트에 값 입력__##########################################
                ws_spec['F1'].value = '결과값 비교'
                for idx, item in enumerate(list_tot_returns):
                    row_num = list_input_rows[idx]
                    mark_data = str(ws_spec['D'+str(row_num)].value)
                    ws_spec['E'+str(row_num)].value = item
                    if mark_data == str(item):
                        ws_spec['F' + str(row_num)].value = 'Match'
                    else:
                        ws_spec['F' + str(row_num)].value = 'Mismatch'
                    self.setPrintText('/s {} 번째 파일 "기본단말_Spec정보" 시트 {}행 데이터 입력 /e'.format(file_idx, row_num))

                ##########################################__spec 시트 E열 옵션 조정__##########################################
                # E column all cell font adjust
                max_row = ws_spec.max_row
                for mCell in ws_spec["E1:E"+str(max_row)]:
                    for cell in mCell:
                        if cell.row == 1:
                            cell.fill = self.gray_fill

                        cell.font = self.red_font
                        cell.alignment = self.general_alignment
                        cell.border = self.thin_border

                # F column all cell font adjust
                max_row = ws_spec.max_row
                for mCell in ws_spec["F1:F"+str(max_row)]:
                    for cell in mCell:
                        if cell.row == 1:
                            cell.fill = self.gray_fill
                            cell.font = self.red_font

                        if cell.value == 'Match':
                            cell.font = self.blue_font
                        elif cell.value == 'Mismatch':
                            cell.font = self.red_font

                        cell.alignment = self.general_alignment
                        cell.border = self.thin_border

                # set filter
                ws_spec.auto_filter.ref = "A1:F"+str(max_row)

                # each column width adjust
                sheet_cell_list = ['A', 'B', 'C', 'D', 'E', 'F']
                sheet_width_list = [21, 31.71, 31.71, 32.43, 32.43, 16.57]
                for i in range(len(sheet_cell_list)):
                    ws_spec.column_dimensions[sheet_cell_list[i]].width = sheet_width_list[i]
                # ws_spec.row_dimensions[1].height = 45

                self.setPrintText('/s {}번 파일 "기본단말_Spec정보" 시트 스타일 적용 완료 /e'.format(file_idx))
                # save file
                wb_output.save(file_out_path)
            else:
                self.setPrintText('/s {} 파일에 필수 sheet가 없습니다! 다시 확인해주세요 /e'.format(file_in_path))
                return
        except:
            self.setPrintText('/s Error: {}. {}, line: {}'.format(sys.exc_info()[0], sys.exc_info()[1], sys.exc_info()[2].tb_lineno)+' /e')
            self.end_count = "y"
            self.end_flag.emit()

    # main method
    def run(self):

        try:
            ###########################__Setting print Text Thread__######################

            self.thread_count = threading.Thread(target=self.getCountRows, args=())
            self.thread_count.daemon = True
            self.thread_count.start()
            self.nowTime = datetime.today().strftime("%Y-%m-%d")

            #################################################################_SETTING INPUT_###########################################################################
            # Save root directory
            self.flag_root = os.path.isdir(self.home+"\\Desktop\\CA\\")
            if not self.flag_root:
                os.mkdir(self.home + "\\Desktop\\CA\\")

            # extract file name each list_files and make every out file path
            for idx, item in enumerate(self.list_files):

                temp_filename = os.path.basename(item)
                temp_filename = re.sub("(.xlsx|.xls)", "", temp_filename)
                output_file = self.home+"\\Desktop\\CA\\"+temp_filename+"_result("+self.nowTime+").xlsx"
                # self.list_out_files.append(output_file)

                #Core Code
                self.start_time = datetime.today().strftime("%Y-%m-%d %H:%M:%S")
                #Excel input Data read
                self.setPrintText("/s STARTED_TIME: "+self.start_time+" /e")
                self.generate_sheet(item, output_file, idx+1)
                self.currentRow = 0
                self.totalRows = self.totalRows + 1

            # job closed
            self.end_time = datetime.today().strftime("%Y-%m-%d %H:%M:%S")
            self.setPrintText("/s FINISHED_TIME: " + self.end_time + " /e")
            self.end_count = "y"
            self.end_flag.emit()

        except:
            self.setPrintText('/s Error: {}. {}, line: {}'.format(sys.exc_info()[0], sys.exc_info()[1], sys.exc_info()[2].tb_lineno)+' /e')
            self.end_count = "y"
            self.end_flag.emit()

if __name__ == '__main__':
    moduler = Formater('C:\\Users\\TestEnC\\Desktop\\VOC\\input_sample.xlsx', 'y',  'f1')
    moduler.run()
